"""
pool_integrator — Merge validated pool files into a unified BMDProject JSON.

After file upload, fingerprinting, and cross-validation, the pool contains
multiple files (xlsx, txt/csv, bm2) covering different endpoint domains
(body_weight, organ_weights, clin_chem, hematology, hormones, gene_expression,
etc.) at different processing tiers.

This module's job is *integration*: select the best file for each domain
(respecting user conflict resolutions), then delegate to bmdx-core's native
Java code for .bm2 merging and JSON serialization.

Data flow:
    Validate → Resolve conflicts → integrate_pool() → integrated.json
        → build_table_data() → section cards with tables & narratives

The Java layer (IntegrateProject) uses BMDExpress 3's native classes:
    - ProjectUtilities.addProjectToProject() for .bm2 merging
    - ExperimentFileUtil.readFile() for .txt/.csv parsing
    - Jackson ObjectMapper with NaN→null serializers for valid JSON output

Python handles:
    - Tier selection (which file to use per domain)
    - xlsx pre-conversion (NTP long-format → pivot)
    - Post-integration metadata enrichment (_meta, _category_lookup)
    - LLM metadata inference (experiment descriptions)

Usage:
    from pool_integrator import integrate_pool

    integrated = integrate_pool(dtxsid, session_dir, fingerprints,
                                coverage_matrix, precedence)
"""

# ---------------------------------------------------------------------------
# Imports
# ---------------------------------------------------------------------------

import csv
import json
import logging
import math
import os
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from experiment_metadata import attach_metadata, infer_experiment_metadata
from file_integrator import (
    _BM2_DOMAIN_MAP,
    _detect_sex_from_filename,
    base_domain,
    detect_domain,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

logger = logging.getLogger(__name__)

# Tier priority for auto-selection: .bm2 preferred (has BMD results),
# then txt/csv (BMDExpress-importable pivot), then xlsx (raw long-format).
# NOTE: xlsx files are the study-team source of truth for doses and animal
# roster, but bm2 files are needed for BMD modeling results.  When both
# exist, bm2 wins the integration (for BMD), and xlsx roster metadata
# is overlaid separately via _collect_xlsx_rosters().
_TIER_PREFERENCE = {"bm2": 1, "txt": 2, "csv": 2, "txt_csv": 2, "xlsx": 3}

# Java helper directory — centralized in java_bridge.py
from java_bridge import JAVA_HELPER_DIR, build_classpath


# ---------------------------------------------------------------------------
# Java integration — call bmdx-core's IntegrateProject
# ---------------------------------------------------------------------------


def _run_integrate_java(
    bm2_paths: list[str],
    txt_paths: list[str],
    output_path: str,
    metadata_path: str | None = None,
) -> None:
    """
    Call IntegrateProject (Java) to merge .bm2 and .txt/.csv files into
    a single BMDProject JSON.

    Uses bmdx-core's native classes:
      - ProjectUtilities.addProjectToProject() for .bm2 merge
      - ExperimentFileUtil.readFile() for .txt/.csv parsing
      - Jackson with NaN→null serializers for valid JSON output

    Args:
        bm2_paths:     List of .bm2 file paths to merge.
        txt_paths:     List of .txt/.csv pivot files to import.
        output_path:   Where to write the integrated JSON.
        metadata_path: Optional path to LLM metadata sidecar JSON.
    """
    from java_bridge import build_classpath

    cp = build_classpath()
    helper_dir = str(JAVA_HELPER_DIR)

    cmd = [
        "java", "-cp", f"{cp}:{helper_dir}",
        "IntegrateProject",
        "--output", output_path,
    ]

    if bm2_paths:
        cmd.append("--bm2")
        cmd.extend(bm2_paths)

    if txt_paths:
        cmd.append("--txt")
        cmd.extend(txt_paths)

    if metadata_path:
        cmd.extend(["--metadata", metadata_path])

    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        logger.error("IntegrateProject failed:\n%s\n%s", result.stdout, result.stderr)
        raise RuntimeError(f"Java integration failed: {result.stderr}")

    logger.info("IntegrateProject output:\n%s", result.stdout)


# ---------------------------------------------------------------------------
# xlsx parser — NTP long-format needs Python pre-conversion
# ---------------------------------------------------------------------------

# Maps base domain names to Apical platform vocabulary values.
# Used to write # Platform: headers on converted wide-format files.
_DOMAIN_TO_PLATFORM: dict[str, str] = {
    "body_weight":    "Body Weight",
    "organ_weights":  "Organ Weight",
    "clin_chem":      "Clinical Chemistry",
    "hematology":     "Hematology",
    "hormones":       "Hormones",
    "tissue_conc":    "Tissue Concentration",
    "clinical_obs":   "Clinical Observations",
}

# Reverse of _BM2_DOMAIN_MAP — maps our canonical domain names to the
# BMDExpress experiment name prefixes.  Both _tox_study and _inferred
# variants are included because either type of file might come through
# xlsx_to_pivot_txt (xlsx files don't have _tox_study in their names,
# so they get _inferred domains from the filename heuristics).
_DOMAIN_TO_BM2_PREFIX: dict[str, str] = {
    "body_weight_tox_study":    "BodyWeight",
    "body_weight_inferred":     "BodyWeight",
    "organ_weights_tox_study":  "OrganWeight",
    "organ_weights_inferred":   "OrganWeight",
    "clin_chem_tox_study":      "ClinicalChemistry",
    "clin_chem_inferred":       "ClinicalChemistry",
    "hematology_tox_study":     "Hematology",
    "hematology_inferred":      "Hematology",
    "hormones_tox_study":       "Hormone",
    "hormones_inferred":        "Hormone",
    "tissue_conc_tox_study":    "TissueConcentration",
    "tissue_conc_inferred":     "TissueConcentration",
    "clinical_obs":             "ClinicalObservation",
}


def _collect_xlsx_rosters(
    coverage_matrix: dict,
    fingerprints: dict,
    files_dir: str,
) -> dict[str, dict]:
    """
    For each domain that has a study xlsx, extract the authoritative animal
    roster (doses, animal IDs, selection groups) from the fingerprint.

    Study xlsx files are the source of truth for which animals were in which
    dose groups — including animals that died before certain endpoints could
    be measured.  This data is stored on integrated._meta.xlsx_rosters so
    that build_table_data() can use it for accurate N counts and footnotes.

    Args:
        coverage_matrix: {domain → {tier_name → [file_ids]}}.
        fingerprints:    {file_id → FileFingerprint or dict}.
        files_dir:       Directory containing uploaded files.

    Returns:
        {domain → {
            "file_id": str,
            "filename": str,
            "dose_groups": [float, ...],
            "n_animals_by_dose": {dose: int},
            "selection_groups": [str, ...],
            "animals_by_dose_selection": {dose: {selection: [animal_ids]}},
        }}
    """
    rosters = {}
    for domain, tiers in coverage_matrix.items():
        xlsx_fids = tiers.get("xlsx")
        if not xlsx_fids:
            continue
        if isinstance(xlsx_fids, str):
            xlsx_fids = [xlsx_fids]
        xlsx_fids = [f for f in xlsx_fids if f]
        if not xlsx_fids:
            continue

        fid = xlsx_fids[0]
        fp = fingerprints.get(fid)
        if fp is None:
            continue

        # Support both dict and FileFingerprint attribute access
        def _get(obj, key, default=None):
            return obj.get(key, default) if isinstance(obj, dict) else getattr(obj, key, default)

        # Only include study files (two-tab NTP structure).
        # For fingerprints from old caches (before is_study_file was added),
        # fall back to assuming all xlsx files are study files — this is safe
        # because non-study xlsx files would never appear in the pool.
        is_study = _get(fp, "is_study_file", None)
        if is_study is None:
            # Legacy fingerprint without the field — assume xlsx = study file
            is_study = _get(fp, "file_type", "") == "xlsx"
        if not is_study:
            continue

        # Stringify float-keyed dicts for JSON serialization safety — the
        # integrated.json is written with json.dump() which rejects non-str keys.
        # Downstream annotate_missing_animals() converts back to float.
        def _str_keys(d):
            if not d:
                return d
            return {str(k): v for k, v in d.items()}

        rosters[domain] = {
            "file_id": fid,
            "filename": _get(fp, "filename", ""),
            "dose_groups": _get(fp, "dose_groups", []),
            "n_animals_by_dose": _str_keys(_get(fp, "n_animals_by_dose", {})),
            "selection_groups": _get(fp, "selection_groups", []),
            "animals_by_dose_selection": _str_keys(_get(fp, "animals_by_dose_selection", {})),
            "core_animals_by_dose_sex": _str_keys(_get(fp, "core_animals_by_dose_sex", {})),
        }

    return rosters


def xlsx_to_pivot_txt(path: str, output_dir: str) -> list[str]:
    """
    Convert an NTP xlsx long-format file into tab-delimited pivot files
    that IntegrateProject can import via ExperimentFileUtil.

    NTP xlsx files are long-format: one row per animal × observation.
    We pivot to BMDExpress format:
        Row 1: column headers (probe_id, colname1, colname2, ...)
        Row 2: doses (label, dose1, dose2, ...)
        Row 3+: endpoint measurements (endpoint_name, value1, value2, ...)

    One output file per sex found in the data.

    Args:
        path:       Absolute path to the xlsx file.
        output_dir: Directory to write the pivot .txt files.

    Returns:
        List of paths to the generated pivot .txt files.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — cannot parse xlsx")
        return []

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        logger.warning("Could not open xlsx %s: %s", path, e)
        return []

    data_sheet = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[-1]]

    # --- Read all rows ---
    headers: list[str] = []
    # {(sex, endpoint): {animal_id: value}}
    data_by_sex_endpoint: dict[tuple[str, str], dict[str, float]] = {}
    animal_dose_map: dict[str, float] = {}
    animal_sex_map: dict[str, str] = {}

    standard_cols = {
        "NTP Study Number", "Concentration", "Animal ID", "Sex",
        "Selection", "Terminal Flag", "",
    }

    for row_idx, row in enumerate(data_sheet.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(c) if c is not None else "" for c in row]
            continue

        if row[1] is None or row[2] is None:
            continue

        try:
            dose = float(row[1])
        except (ValueError, TypeError):
            continue

        animal_id = str(row[2]).strip()
        sex = str(row[3]).strip().title() if row[3] is not None else "Unknown"
        animal_dose_map[animal_id] = dose
        animal_sex_map[animal_id] = sex

        for col_idx in range(4, len(headers)):
            col_name = headers[col_idx].strip()
            if col_name in standard_cols or not col_name:
                continue

            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue

            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue

            key = (sex, col_name)
            data_by_sex_endpoint.setdefault(key, {})[animal_id] = fval

    wb.close()

    if not data_by_sex_endpoint:
        return []

    # --- Detect domain from filename ---
    filename = os.path.basename(path)
    file_size = os.path.getsize(path) if os.path.exists(path) else 0
    domain = detect_domain(filename, "xlsx", file_size)
    prefix = _DOMAIN_TO_BM2_PREFIX.get(domain, "Unknown")

    # --- Group by sex and write pivot files ---
    # Collect all endpoints and animals per sex
    sex_data: dict[str, dict] = {}
    for (sex, endpoint), animal_vals in data_by_sex_endpoint.items():
        if sex not in sex_data:
            sex_data[sex] = {"endpoints": {}, "animals": set()}
        sex_data[sex]["endpoints"][endpoint] = animal_vals
        sex_data[sex]["animals"].update(animal_vals.keys())

    output_paths = []
    for sex, info in sex_data.items():
        # Sort animals by dose then ID for consistent column order
        animals = sorted(info["animals"],
                         key=lambda a: (animal_dose_map.get(a, 0), a))

        # Build the wide-format file.
        # Metadata headers are parsed by ExperimentDescriptionParser so that
        # ExperimentDescription fields are set during Java import.
        lines = []
        # Metadata headers — provider, platform, data type
        lines.append(f"# Provider: Apical")
        lines.append(f"# Platform: {_DOMAIN_TO_PLATFORM.get(base_domain(domain), 'Generic')}")
        lines.append(f"# Data Type: {'tox_study' if domain.endswith('_tox_study') else 'inferred'}")
        # Row 1: header — probe_id + animal IDs
        lines.append("\t".join([f"{prefix}{sex}"] + animals))
        # Row 2: doses
        lines.append("\t".join(["Dose"] + [
            str(animal_dose_map.get(a, 0)) for a in animals
        ]))
        # Row 3+: endpoints
        for endpoint, animal_vals in info["endpoints"].items():
            values = []
            for a in animals:
                v = animal_vals.get(a)
                values.append(str(v) if v is not None else "")
            lines.append("\t".join([endpoint] + values))

        # Write to file
        safe_sex = sex.replace(" ", "_")
        out_name = f"{prefix}{safe_sex}.txt"
        out_path = os.path.join(output_dir, out_name)
        with open(out_path, "w") as f:
            f.write("\n".join(lines) + "\n")

        output_paths.append(out_path)
        logger.info("Converted long-format xlsx → wide: %s (%d endpoints, %d animals)",
                     out_name, len(info["endpoints"]), len(animals))

    return output_paths


def tox_study_csv_to_pivot_txt(
    path: str,
    output_dir: str,
    domain: str,
) -> list[str]:
    """
    Convert a long-format _tox_study CSV file into tab-delimited pivot files
    that IntegrateProject can import via ExperimentFileUtil.

    _tox_study CSV files are long-format like NTP xlsx files:
      Concentration, Animal ID, Sex, endpoint1, endpoint2, ...

    The output pivot format is the same as xlsx_to_pivot_txt():
      Row 0: # Domain: <domain>   (metadata for ExperimentDescriptionParser)
      Row 1: experiment_name, animal_id1, animal_id2, ...
      Row 2: Dose, dose1, dose2, ...
      Row 3+: endpoint_name, value1, value2, ...

    One output file per sex found in the data.

    Args:
        path:       Absolute path to the _tox_study CSV file.
        output_dir: Directory to write the pivot .txt files.
        domain:     The canonical domain (e.g., "body_weight_tox_study").

    Returns:
        List of paths to the generated pivot .txt files.
    """
    # Detect separator (CSV or TSV)
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        first_line = f.readline()
    sep = "\t" if "\t" in first_line else ","

    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f, delimiter=sep)
        rows = list(reader)

    if len(rows) < 2:
        logger.warning("tox_study CSV too short: %s", path)
        return []

    header = [c.strip() for c in rows[0]]
    header_lower = [h.lower() for h in header]

    # Find key column indices — same structure as NTP xlsx
    _LONG_FORMAT_MARKERS = {"concentration", "animal id", "sex"}
    if not _LONG_FORMAT_MARKERS.issubset(set(header_lower)):
        logger.warning("tox_study CSV missing required columns: %s", path)
        return []

    dose_col = header_lower.index("concentration")
    animal_col = header_lower.index("animal id")
    sex_col = header_lower.index("sex") if "sex" in header_lower else None

    # Endpoint columns: everything not in known metadata columns
    _META_COLS = {
        "ntp study number", "concentration", "animal id", "sex",
        "selection", "observation day", "terminal flag",
        "observation", "site", "modifier", "severity",
    }
    endpoint_cols = [
        i for i, h in enumerate(header_lower)
        if h not in _META_COLS and h
    ]

    if not endpoint_cols:
        logger.warning("No endpoint columns found in %s", path)
        return []

    # Parse data rows into {(sex, endpoint): {animal_id: value}}
    data_by_sex_endpoint: dict[tuple[str, str], dict[str, float]] = {}
    animal_dose_map: dict[str, float] = {}

    for row in rows[1:]:
        if len(row) <= max(dose_col, animal_col):
            continue

        aid = row[animal_col].strip()
        if not aid:
            continue

        try:
            dose = float(row[dose_col].strip())
        except (ValueError, TypeError):
            continue

        animal_dose_map[aid] = dose

        sex = "Unknown"
        if sex_col is not None and sex_col < len(row):
            s = row[sex_col].strip()
            if s:
                sex = s.title()

        for col_idx in endpoint_cols:
            col_name = header[col_idx]
            if col_idx >= len(row):
                continue
            val = row[col_idx].strip()
            if not val:
                continue
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue

            key = (sex, col_name)
            data_by_sex_endpoint.setdefault(key, {})[aid] = fval

    if not data_by_sex_endpoint:
        return []

    # Build BMDExpress experiment name prefix from domain
    prefix = _DOMAIN_TO_BM2_PREFIX.get(domain, base_domain(domain).title())

    # Group by sex and write pivot files
    sex_data: dict[str, dict] = {}
    for (sex, endpoint), animal_vals in data_by_sex_endpoint.items():
        if sex not in sex_data:
            sex_data[sex] = {"endpoints": {}, "animals": set()}
        sex_data[sex]["endpoints"][endpoint] = animal_vals
        sex_data[sex]["animals"].update(animal_vals.keys())

    output_paths = []
    for sex, info in sex_data.items():
        animals = sorted(info["animals"],
                         key=lambda a: (animal_dose_map.get(a, 0), a))

        lines = []
        # Metadata headers for ExperimentDescriptionParser
        lines.append(f"# Provider: Apical")
        lines.append(f"# Platform: {_DOMAIN_TO_PLATFORM.get(base_domain(domain), 'Generic')}")
        lines.append(f"# Data Type: {'tox_study' if domain.endswith('_tox_study') else 'inferred'}")
        # Row 1: header — probe_id + animal IDs
        lines.append("\t".join([f"{prefix}{sex}"] + animals))
        # Row 2: doses
        lines.append("\t".join(["Dose"] + [
            str(animal_dose_map.get(a, 0)) for a in animals
        ]))
        # Row 3+: endpoints
        for endpoint, animal_vals in info["endpoints"].items():
            values = [str(animal_vals.get(a, "")) for a in animals]
            lines.append("\t".join([endpoint] + values))

        safe_sex = sex.replace(" ", "_")
        out_name = f"{prefix}{safe_sex}.txt"
        out_path = os.path.join(output_dir, out_name)
        with open(out_path, "w") as f:
            f.write("\n".join(lines) + "\n")

        output_paths.append(out_path)
        logger.info("Converted long-format CSV → wide: %s (%d endpoints, %d animals)",
                     out_name, len(info["endpoints"]), len(animals))

    return output_paths


# ---------------------------------------------------------------------------
# Post-integration domain stamping
# ---------------------------------------------------------------------------


def _stamp_domains(
    integrated: dict,
    source_files: dict[str, dict],
    fingerprints: dict[str, dict],
) -> None:
    """
    Set domain on each experiment's experimentDescription using authoritative
    fingerprint data, overriding any LLM-inferred domain.

    The source_files dict maps domain → {filename, file_id, tier}.  For each
    experiment, we determine which source file it came from (by matching
    experiment names to source filenames and _BM2_DOMAIN_MAP prefixes) and
    set the domain accordingly.

    This runs AFTER LLM metadata inference and provides an authoritative
    domain assignment that doesn't depend on LLM accuracy.

    Args:
        integrated:   The merged BMDProject dict (mutated in place).
        source_files: domain → {filename, file_id, tier} from integrate_pool.
        fingerprints: file_id → fingerprint dict for all pool files.
    """
    # Build reverse map: experiment name prefix → domain.
    # For each source file, determine what experiment names it would produce.
    # bm2 files: experiments keep their original names (mapped via _BM2_DOMAIN_MAP)
    # txt files: experiment name = first cell of header row (filename-derived)
    # xlsx-converted: experiment name = {prefix}{sex} (from _DOMAIN_TO_BM2_PREFIX)

    # First, build a map from file_id → domain for quick lookup
    fid_to_domain: dict[str, str] = {}
    for domain, info in source_files.items():
        fid = info.get("file_id")
        if fid:
            fid_to_domain[fid] = domain

    # Build experiment-name → domain mapping from source filenames
    # For txt/csv files, the experiment name is typically the filename stem
    # or the first header cell. For xlsx-converted files, it's {prefix}{sex}.
    filename_to_domain: dict[str, str] = {}
    for domain, info in source_files.items():
        fname = info.get("filename", "")
        if fname:
            # Store both with and without extension for matching
            filename_to_domain[fname] = domain
            stem = os.path.splitext(fname)[0]
            filename_to_domain[stem] = domain

    experiments = integrated.get("doseResponseExperiments", [])
    stamped = 0

    for exp in experiments:
        exp_name = exp.get("name", "")
        desc = exp.get("experimentDescription")
        if desc is None:
            desc = {}
            exp["experimentDescription"] = desc

        # If Java already set domain (from # Domain: header), trust it.
        if desc.get("domain"):
            stamped += 1
            continue

        matched_domain = None

        # Strategy 1: Direct experiment name → _BM2_DOMAIN_MAP prefix match.
        # Strip underscores, spaces, and sex prefix before matching — BMDExpress
        # experiment names can have sex as prefix ("female_organ_weights") or
        # suffix ("OrganWeightFemale"), and _BM2_DOMAIN_MAP expects just the
        # domain prefix ("organweight").
        name_lower = exp_name.lower().replace("_", "").replace(" ", "")
        # Strip "female"/"male" — strip "female" FIRST since it contains "male"
        name_stripped = name_lower.replace("female", "").replace("male", "").strip()
        for prefix, bm2_domain in _BM2_DOMAIN_MAP.items():
            if name_stripped.startswith(prefix) or name_lower.startswith(prefix):
                if bm2_domain in source_files:
                    matched_domain = bm2_domain
                    break

        # Strategy 2: Match experiment name to source filenames
        if not matched_domain:
            for fname, dom in filename_to_domain.items():
                stem = os.path.splitext(fname)[0]
                # Experiment name might be the filename stem or derived from it
                if exp_name == stem or exp_name.startswith(stem):
                    matched_domain = dom
                    break

        # Strategy 3: Use _DOMAIN_TO_BM2_PREFIX reverse lookup.
        # Case-insensitive, also tries after stripping sex prefix.
        if not matched_domain:
            for dom, prefix in _DOMAIN_TO_BM2_PREFIX.items():
                prefix_lower = prefix.lower()
                if (name_lower.startswith(prefix_lower) or
                    name_stripped.startswith(prefix_lower)):
                    if dom in source_files:
                        matched_domain = dom
                        break

        if matched_domain:
            desc["domain"] = matched_domain
            stamped += 1

    logger.info("Stamped domain on %d/%d experiments", stamped, len(experiments))


# ---------------------------------------------------------------------------
# Main integration orchestrator
# ---------------------------------------------------------------------------


def integrate_pool(
    dtxsid: str,
    session_dir: str,
    fingerprints: dict[str, dict],
    coverage_matrix: dict[str, dict[str, list[str]]],
    precedence: list[dict],
    *,
    test_article: dict | None = None,
    llm_generate_json: Any = None,
) -> dict:
    """
    Merge all pool files into a single unified BMDProject JSON.

    For each domain in the coverage matrix:
      1. If the user resolved a conflict → use that chosen file.
      2. Otherwise, prefer .bm2 (has BMD results), then txt/csv, then xlsx.
      3. Collect .bm2 and .txt/.csv file paths for Java integration.
      4. Convert .xlsx files to pivot .txt format for Java import.
      5. Call IntegrateProject (Java) to merge everything with native
         BMDExpress classes and Jackson NaN-safe serialization.
      6. Read back the JSON, add _meta and _category_lookup for the
         process-integrated endpoint.

    Args:
        dtxsid:            The DTXSID identifying this session.
        session_dir:       Absolute path to sessions/{dtxsid}/.
        fingerprints:      Dict mapping file_id → fingerprint dict.
        coverage_matrix:   Dict mapping domain → tier_name → list[file_id].
        precedence:        List of user conflict resolutions.
        test_article:      Dict with 'name', 'casrn', 'dsstox' for metadata.
        llm_generate_json: Callable for LLM metadata inference.

    Returns:
        The merged BMDProject dict, also saved to sessions/{dtxsid}/integrated.json.
    """
    files_dir = os.path.join(session_dir, "files")
    out_path = os.path.join(session_dir, "integrated.json")

    # Collect user-resolved file IDs for quick lookup
    resolved_file_ids: dict[str, str] = {}
    for entry in precedence:
        fid = entry.get("chosen_file_id", "")
        if not fid:
            continue
        fp = fingerprints.get(fid, {})
        domain = fp.get("domain") if isinstance(fp, dict) else getattr(fp, "domain", None)
        if domain:
            resolved_file_ids[domain] = fid

    # --- Select files per domain ---
    # Separate into lists for Java integration
    bm2_paths: list[str] = []
    txt_paths: list[str] = []
    xlsx_paths: list[str] = []
    tox_study_csv_paths: list[tuple[str, str]] = []  # (path, domain)
    source_files: dict[str, dict] = {}

    for domain, tiers in coverage_matrix.items():
        chosen_fid = None
        chosen_tier = None

        # 1. Check user conflict resolution
        if domain in resolved_file_ids:
            chosen_fid = resolved_file_ids[domain]
            fp = fingerprints.get(chosen_fid, {})
            chosen_tier = (
                fp.get("file_type") if isinstance(fp, dict)
                else getattr(fp, "file_type", None)
            )

        # 2. Auto-select by tier preference: bm2 > txt/csv > xlsx.
        # Include ALL files of the winning tier (e.g., both male and female
        # CSVs), not just the first — each sex is a separate file.
        chosen_fids: list[str] = []
        if not chosen_fid:
            best_priority = 999
            for tier_name, fids_raw in tiers.items():
                if fids_raw is None:
                    continue
                fids = fids_raw if isinstance(fids_raw, list) else [fids_raw]
                fids = [f for f in fids if f]
                priority = _TIER_PREFERENCE.get(tier_name, 999)
                if priority < best_priority and fids:
                    best_priority = priority
                    chosen_fids = fids
                    chosen_fid = fids[0]
                    chosen_tier = tier_name
        else:
            chosen_fids = [chosen_fid]

        if not chosen_fid:
            logger.warning("No file found for domain %s — skipping", domain)
            continue

        # Route ALL files of the winning tier to the appropriate list
        first_filename = None
        for fid in chosen_fids:
            fp = fingerprints.get(fid, {})
            filename = fp.get("filename") if isinstance(fp, dict) else getattr(fp, "filename", "")
            actual_file_type = (
                fp.get("file_type") if isinstance(fp, dict)
                else getattr(fp, "file_type", chosen_tier)
            )
            file_path = os.path.join(files_dir, filename) if filename else ""

            if not file_path or not os.path.exists(file_path):
                logger.warning("File not found for domain %s (id=%s) — skipping", domain, fid)
                continue

            if first_filename is None:
                first_filename = filename

            logger.info("Selected domain=%s from %s (%s)", domain, filename, actual_file_type)

            # Route to the appropriate list for Java integration
            if chosen_tier == "bm2" or actual_file_type == "bm2":
                bm2_paths.append(file_path)
            elif actual_file_type in ("txt", "csv") or chosen_tier == "txt_csv":
                # Gene expression txt/csv are raw matrices, not clinical pivots —
                # they can't be imported directly.  Skip them; gene expression
                # comes from the .bm2 tier via BMDExpress's analysis pipeline.
                if domain == "gene_expression":
                    continue
                # _tox_study CSV files are long-format (like xlsx) and need
                # conversion to pivot format.  Queue them for conversion.
                if domain.endswith("_tox_study") and actual_file_type == "csv":
                    tox_study_csv_paths.append((file_path, domain))
                else:
                    txt_paths.append(file_path)
            elif actual_file_type == "xlsx" or chosen_tier == "xlsx":
                xlsx_paths.append(file_path)
            else:
                logger.warning("Unknown tier %s for domain %s — skipping", chosen_tier, domain)

        # Record source file metadata (use first file as representative)
        if first_filename:
            source_files[domain] = {
                "file_id": chosen_fids[0], "filename": first_filename,
                "tier": chosen_tier or actual_file_type,
                "file_count": len(chosen_fids),
            }

    # --- Convert long-format files to pivot format for Java import ---
    # NTP xlsx files and _tox_study CSV files are long-format (one row
    # per animal).  ExperimentFileUtil expects tab-delimited pivot format.
    pivot_dir = os.path.join(session_dir, "_pivots")
    if xlsx_paths or tox_study_csv_paths:
        os.makedirs(pivot_dir, exist_ok=True)
        for xlsx_path in xlsx_paths:
            pivot_files = xlsx_to_pivot_txt(xlsx_path, pivot_dir)
            txt_paths.extend(pivot_files)
        for csv_path, csv_domain in tox_study_csv_paths:
            pivot_files = tox_study_csv_to_pivot_txt(csv_path, pivot_dir, csv_domain)
            txt_paths.extend(pivot_files)

    # --- Call Java to merge everything ---
    if not bm2_paths and not txt_paths:
        logger.warning("No files to integrate for %s", dtxsid)
        return {}

    _run_integrate_java(bm2_paths, txt_paths, out_path)

    # --- Read back the Java-produced JSON ---
    # IntegrateProject writes valid JSON (NaN→null) via Jackson.
    # We add our metadata overlay (_meta, _category_lookup) for the
    # process-integrated endpoint.
    with open(out_path, "r") as f:
        integrated = json.load(f)

    # --- Collect xlsx animal rosters for all domains ---
    # Even when bm2 wins the integration (for BMD results), the study xlsx
    # is authoritative for doses and animal roster.  Overlay this so
    # build_table_data() can use xlsx N counts and detect dead animals.
    xlsx_rosters = _collect_xlsx_rosters(coverage_matrix, fingerprints, files_dir)

    # Add provenance metadata
    integrated["_meta"] = {
        "dtxsid": dtxsid,
        "integrated_at": datetime.now(tz=timezone.utc).isoformat(),
        "source_files": source_files,
        "xlsx_rosters": xlsx_rosters,
    }

    # --- Build category lookup via native Java API ---
    # ExportCategories.java reads CategoryAnalysisResults directly from the
    # BMDProject — no DataCombinerService, no TSV.  This replaces the broken
    # TSV export path that crashed with IndexOutOfBoundsException.
    cat_results = integrated.get("categoryAnalysisResults", [])
    if cat_results:
        from apical_report import export_categories, build_category_lookup

        # Collect all experiment names from the integrated data so
        # build_category_lookup can resolve BMDExpress pipeline suffixes
        # (e.g., "_williams_0.05_NOMTC_nofoldfilter") back to the raw
        # experiment names used by build_table_data().
        all_exp_names = [
            exp.get("name", "")
            for exp in integrated.get("doseResponseExperiments", [])
        ]

        merged_cat_lookup: dict[tuple[str, str], dict] = {}
        for bm2_path in bm2_paths:
            cat_json_path = os.path.join(
                session_dir, f"_cat_{os.path.basename(bm2_path)}.json",
            )
            try:
                categories_json = export_categories(bm2_path, cat_json_path)
                cat_lookup = build_category_lookup(
                    categories_json, experiment_names=all_exp_names,
                )
                merged_cat_lookup.update(cat_lookup)
            except Exception as e:
                logger.warning("Category export failed for %s: %s", bm2_path, e)
            finally:
                if os.path.exists(cat_json_path):
                    os.remove(cat_json_path)

        integrated["_category_lookup"] = {
            f"{k[0]}|{k[1]}": v
            for k, v in merged_cat_lookup.items()
        }
    else:
        integrated["_category_lookup"] = {}

    # --- Infer structured ExperimentDescription metadata via LLM ---
    if test_article and llm_generate_json:
        all_experiments = integrated.get("doseResponseExperiments", [])
        descriptions = infer_experiment_metadata(
            experiments=all_experiments,
            source_files=source_files,
            test_article=test_article,
            llm_generate_json=llm_generate_json,
        )
        attach_metadata(all_experiments, descriptions)

    # --- Authoritative domain stamping from fingerprint data ---
    # Override LLM-inferred domain with the authoritative domain from
    # source_files (which comes directly from file fingerprinting).
    _stamp_domains(integrated, source_files, fingerprints)

    # Re-write the JSON with our metadata additions
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(integrated, f, indent=2)

    # --- Write enriched .bm2 file ---
    # Serialize the metadata-enriched BMDProject back to the canonical .bm2
    # format (Java ObjectOutputStream).  This closes the metadata loop:
    # the LLM inferred metadata, the user will approve it, and the .bm2
    # becomes the single source of truth — openable in BMDExpress 3 with
    # all ExperimentDescription fields pre-populated.
    #
    # The .bm2 is written alongside integrated.json in the session directory.
    # It can be re-exported on demand via export_integrated_bm2().
    bm2_out = os.path.join(session_dir, "integrated.bm2")
    try:
        export_integrated_bm2(out_path, bm2_out)
    except Exception as e:
        # Non-fatal: the JSON is the primary artifact; .bm2 is a convenience.
        logger.warning("Failed to write enriched .bm2: %s", e)

    logger.info(
        "Integration complete: %d experiments, %d BMD results, %d category results → %s",
        len(integrated.get("doseResponseExperiments", [])),
        len(integrated.get("bMDResult", [])),
        len(integrated.get("categoryAnalysisResults", [])),
        out_path,
    )

    return integrated


def export_integrated_bm2(json_path: str, bm2_path: str) -> str:
    """
    Convert a metadata-enriched integrated.json back to .bm2 format.

    Uses JsonToBm2.java to deserialize the Jackson-annotated JSON into a
    BMDProject object graph, then write it via ObjectOutputStream.  The
    resulting .bm2 file is a standard BMDExpress 3 project file that can
    be opened natively — with ExperimentDescription metadata pre-filled.

    Args:
        json_path: Path to the integrated.json (with @type/@ref annotations).
        bm2_path:  Path to write the output .bm2 file.

    Returns:
        The bm2_path on success.

    Raises:
        RuntimeError: If the Java process exits with a non-zero code.
    """
    classpath = build_classpath()
    cmd = [
        "java", "-cp", classpath,
        "JsonToBm2", json_path, bm2_path,
    ]
    result = subprocess.run(
        cmd, capture_output=True, text=True, timeout=30,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"JsonToBm2 failed (exit {result.returncode}): {result.stderr}"
        )
    logger.info("Wrote enriched .bm2: %s", bm2_path)
    return bm2_path
