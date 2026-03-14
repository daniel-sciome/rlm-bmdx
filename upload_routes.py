"""
upload_routes.py — File upload and processing API endpoints.

Extracted from background_server.py.  These endpoints handle uploading,
processing, and previewing .bm2, .csv, .txt, .xlsx, and .zip files for
the 5dToxReport file pool.

All endpoints are mounted as a FastAPI APIRouter on the /api prefix.

Endpoints:
  POST /api/upload-bm2       — Upload .bm2 files for apical endpoint analysis
  POST /api/upload-csv       — Upload gene-level BMD CSV for transcriptomic analysis
  POST /api/upload-zip       — Upload a .zip archive; extract and register files
  POST /api/process-bm2      — Process a .bm2 file and return table data as JSON
  POST /api/process-genomics  — Rank GO gene sets and genes by BMD from CSV data
  GET  /api/preview/{file_id} — Preview uploaded file contents
"""

import asyncio
import json
import logging
import os
import shutil
import tempfile
import uuid
import zipfile
from pathlib import Path

import orjson
from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import JSONResponse, Response

import bm2_cache
from session_store import session_dir
from pool_orchestrator import (
    fingerprint_and_store, run_lightweight_validation, serialize_table_rows,
)
from server_state import (
    get_bm2_uploads,
    get_csv_uploads,
    get_data_uploads,
)
from apical_report import (
    build_table_data_from_bm2,
    generate_results_narrative,
)
from interpret import (
    ToxKBQuerier,
    fetch_gene_descriptions,
    fetch_go_descriptions,
    load_dose_response,
    rank_go_sets_by_bmd,
    rank_genes_by_bmd,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Valid extensions that we extract from zip archives.  Anything else is
# silently skipped to avoid polluting the file pool with irrelevant files.
_ZIP_VALID_EXTENSIONS = {".bm2", ".csv", ".txt", ".xlsx"}


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------

router = APIRouter()


# ---------------------------------------------------------------------------
# POST /api/upload-bm2 — upload .bm2 files for apical endpoint analysis
# ---------------------------------------------------------------------------

@router.post("/api/upload-bm2")
async def api_upload_bm2(request: Request, files: list[UploadFile] = File(...)):
    """
    Accept one or more .bm2 file uploads.

    Saves each file to the session's files/ directory (if a dtxsid query
    parameter is provided) so it persists across page reloads, or to a
    temp directory as a fallback.  Returns a JSON list of metadata objects:
    {id, filename}.  The id is a UUID that other endpoints use to
    reference the uploaded file.
    """
    _bm2_uploads = get_bm2_uploads()

    # If the client provides a DTXSID, save directly to the session's
    # files/ directory so the file survives page reloads and server
    # restarts — no need to wait for approve to persist.
    dtxsid = request.query_params.get("dtxsid", "")
    persist_dir = None
    if dtxsid:
        persist_dir = session_dir(dtxsid) / "files"
        persist_dir.mkdir(exist_ok=True)

    results = []

    for upload in files:
        bm2_id = str(uuid.uuid4())
        safe_name = os.path.basename(upload.filename or "upload.bm2")
        content = await upload.read()

        if persist_dir:
            # Save to the session's files/ directory for persistence
            file_path = str(persist_dir / safe_name)
            with open(file_path, "wb") as f:
                f.write(content)
        else:
            # Fallback: save to a temp directory (lost on restart)
            tmp_dir = tempfile.mkdtemp(prefix="bm2_")
            file_path = os.path.join(tmp_dir, safe_name)
            with open(file_path, "wb") as f:
                f.write(content)

        _bm2_uploads[bm2_id] = {
            "filename": safe_name,
            "temp_path": file_path,
            "table_data": None,   # populated by /api/process-bm2 or preview
            "bm2_json": None,     # populated on first preview/process
            "narrative": None,    # populated by /api/process-bm2
        }

        # Fingerprint the uploaded file for cross-validation.
        # bm2 fingerprints are partial at this stage (no JSON yet) because
        # we haven't run the expensive Java export — domain is detected
        # from the filename only.  Full metadata fills in on first preview.
        validation_issues = []
        if dtxsid:
            fp = fingerprint_and_store(bm2_id, safe_name, file_path, "bm2", dtxsid)
            validation_issues = run_lightweight_validation(fp, dtxsid)

        results.append({
            "id": bm2_id,
            "filename": safe_name,
            "validation_issues": validation_issues,
        })

    return JSONResponse(results)


# ---------------------------------------------------------------------------
# POST /api/process-bm2 — analyze a .bm2 file and return table data as JSON
# ---------------------------------------------------------------------------

@router.post("/api/process-bm2")
async def api_process_bm2(request: Request):
    """
    Process an uploaded .bm2 file through the NTP statistical pipeline.

    Input JSON:
      {
        "bm2_id": "<uuid from upload>",
        "section_title": "Animal Condition, Body Weights, and Organ Weights",
        "table_caption_template": "Summary of ... {sex} ... {compound} ...",
        "compound_name": "PFHxSAm",
        "dose_unit": "mg/kg"
      }

    Calls build_table_data_from_bm2() from apical_report.py, which exports
    the .bm2 via the BMDExpress Java CLI and runs Jonckheere → Williams/Dunnett
    tests.

    Returns JSON with the table data for HTML preview:
      {
        "bm2_id": "...",
        "tables": {
          "Male": [
            {
              "label": "Terminal Body Wt.",
              "doses": [0.0, 0.15, 0.5, 1.5, 5.0],
              "values": {"0.0": "330.2 ± 5.1", ...},
              "n": {"0.0": 5, ...},
              "bmd": "8.49", "bmdl": "4.23", "trend_marker": "**"
            }, ...
          ],
          "Female": [...]
        }
      }
    """
    _bm2_uploads = get_bm2_uploads()

    body = await request.json()
    bm2_id = body.get("bm2_id", "")
    compound_name = body.get("compound_name", "Test Compound")
    dose_unit = body.get("dose_unit", "mg/kg")

    upload = _bm2_uploads.get(bm2_id)
    if not upload:
        return JSONResponse(
            {"error": f"Unknown bm2_id: {bm2_id}"},
            status_code=404,
        )

    bm2_path = upload["temp_path"]
    if not os.path.exists(bm2_path):
        return JSONResponse(
            {"error": f"Uploaded file no longer exists: {upload['filename']}"},
            status_code=410,
        )

    try:
        # Run the full export-and-analyze pipeline in a thread pool
        # because it spawns Java subprocesses (blocking I/O)
        loop = asyncio.get_running_loop()
        table_data, _cat_lookup, bm2_json = await loop.run_in_executor(
            None, build_table_data_from_bm2, bm2_path,
        )

        # Cache the table_data and full JSON so /api/export-docx and
        # /api/preview can reuse them without re-running the Java export
        upload["table_data"] = table_data
        upload["bm2_json"] = bm2_json

        # Generate the NTP-style results narrative from the table data.
        # This produces paragraphs describing body weight and organ weight
        # findings that the user can edit in the UI before export.
        narrative = generate_results_narrative(
            table_data, compound_name, dose_unit,
        )
        upload["narrative"] = narrative

        # Serialize TableRow objects to JSON-friendly dicts
        tables_json = serialize_table_rows(table_data)

        return JSONResponse({
            "bm2_id": bm2_id,
            "tables": tables_json,
            "narrative": narrative,  # list of auto-generated paragraph strings
        })

    except Exception as e:
        return JSONResponse(
            {"error": f"Processing failed: {e}"},
            status_code=500,
        )


# ---------------------------------------------------------------------------
# POST /api/upload-csv — upload loose CSV/TXT data files into the data pool
# ---------------------------------------------------------------------------
# CSV and TXT files are format-equivalent (comma vs tab separator) and can
# contain either apical endpoint data (body weights, organ weights, etc.) or
# gene-level BMD results (genomics).  We don't know which until the pool
# integrator fingerprints them, so we treat all loose CSVs the same as
# CSVs extracted from zip archives: register in _data_uploads, fingerprint,
# and let the pool pipeline determine the domain.

@router.post("/api/upload-csv")
async def api_upload_csv(
    request: Request,
    files: list[UploadFile] = File(...),
):
    """
    Accept one or more loose CSV (or TXT) file uploads.

    Unlike .bm2 files which have a dedicated binary format, CSV/TXT files
    can be either apical endpoint data (tab/comma separated dose-response
    tables) or gene-level BMD output.  The file is saved and registered
    in the generic _data_uploads store — the same store used for CSV/TXT
    files extracted from zip archives.  The pool integrator will
    fingerprint them to determine the domain (body_weight, hormones,
    gene_expression, etc.) during validation.

    Accepts an optional ?dtxsid= query parameter.  When provided, files
    are saved to sessions/{dtxsid}/files/ for persistence across restarts
    and fingerprinting is run immediately.

    Returns JSON:
      [{id, filename, type, validation_issues}, ...]
    """
    _data_uploads = get_data_uploads()

    # Optional DTXSID for persistence and fingerprinting
    dtxsid = request.query_params.get("dtxsid", "")

    results = []

    for upload in files:
        file_id = str(uuid.uuid4())
        safe_name = os.path.basename(upload.filename or "upload.csv")
        ext = os.path.splitext(safe_name)[1].lower()
        file_type = ext.lstrip(".")  # "csv" or "txt"

        # Persist to session directory if DTXSID is available,
        # otherwise save to a temp directory (lost on restart).
        if dtxsid:
            d = session_dir(dtxsid) / "files"
            d.mkdir(parents=True, exist_ok=True)
            file_path = str(d / safe_name)
        else:
            tmp_dir = tempfile.mkdtemp(prefix=f"{file_type}_")
            file_path = os.path.join(tmp_dir, safe_name)

        with open(file_path, "wb") as f:
            content = await upload.read()
            f.write(content)

        # Register in the generic data pool — same store as zip-extracted
        # CSV/TXT files.  The pool integrator fingerprints this file during
        # validation to determine its domain (apical vs genomics).
        _data_uploads[file_id] = {
            "filename": safe_name,
            "temp_path": file_path,
            "type": file_type,
        }

        # Fingerprint for cross-validation if DTXSID is available
        v_issues = []
        if dtxsid:
            fp = fingerprint_and_store(
                file_id, safe_name, file_path, file_type, dtxsid,
            )
            v_issues = run_lightweight_validation(fp, dtxsid)

        results.append({
            "id": file_id,
            "filename": safe_name,
            "type": file_type,
            "validation_issues": v_issues,
        })

    return JSONResponse(results)


# ---------------------------------------------------------------------------
# POST /api/upload-zip — extract a .zip archive and register individual files
# ---------------------------------------------------------------------------
# Accepts a single .zip file, extracts its contents to temp directories,
# and registers each file with the appropriate in-memory store (_bm2_uploads
# or _csv_uploads) just as if they had been uploaded individually via
# /api/upload-bm2 or /api/upload-csv.
#
# Supported file types inside the zip:
#   .bm2  — registered as bm2 uploads (same as /api/upload-bm2)
#   .csv  — parsed via load_dose_response() and registered as csv uploads
#   .txt  — registered as generic text files (stored in uploadedFiles on client)
#   .xlsx — registered as generic spreadsheet files (stored in uploadedFiles)
#
# Files in subdirectories within the zip are extracted with their basename
# only (flattened).  __MACOSX/ entries and hidden files (starting with .)
# are silently skipped.

@router.post("/api/upload-zip")
async def api_upload_zip(request: Request, file: UploadFile = File(...)):
    """
    Accept a single .zip archive upload, extract its contents, and register
    each recognized file type as if it had been uploaded individually.

    .bm2 files are BMDExpress output — they go into _bm2_uploads for
    downstream processing via /api/process-bm2.

    .csv, .txt, .xlsx are raw dose-response experimental data (animal IDs
    across columns, dose concentrations in row 2, measured endpoints in
    subsequent rows — the format BMDExpress imports as doseResponseExperiment).
    These are stored as generic file pool entries.

    Returns a JSON object with:
      bm2_files:   [{id, filename}, ...]            — same shape as /api/upload-bm2
      other_files: [{id, filename, type}, ...]       — csv, txt, and xlsx raw data
      skipped:     [filename, ...]                   — files with unrecognized extensions
    """
    _bm2_uploads = get_bm2_uploads()
    _data_uploads = get_data_uploads()

    # Save the uploaded zip to a temp file so zipfile can read it
    zip_tmp_dir = tempfile.mkdtemp(prefix="zip_")
    zip_path = os.path.join(zip_tmp_dir, "upload.zip")

    with open(zip_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # Validate that it's actually a zip file
    if not zipfile.is_zipfile(zip_path):
        shutil.rmtree(zip_tmp_dir, ignore_errors=True)
        return JSONResponse(
            {"error": "The uploaded file is not a valid .zip archive"},
            status_code=400,
        )

    # If the client provides a DTXSID, persist .bm2 files to the session's
    # files/ directory immediately so they survive page reloads.
    dtxsid = request.query_params.get("dtxsid", "")
    persist_dir = None
    if dtxsid:
        persist_dir = session_dir(dtxsid) / "files"
        persist_dir.mkdir(exist_ok=True)

    bm2_results = []
    other_results = []  # csv, txt, xlsx — raw data files
    skipped = []

    with zipfile.ZipFile(zip_path, "r") as zf:
        for member in zf.namelist():
            # Skip directories, __MACOSX metadata, and hidden files
            if member.endswith("/"):
                continue
            basename = os.path.basename(member)
            if not basename or basename.startswith(".") or "__MACOSX" in member:
                continue

            # Check file extension
            _, ext = os.path.splitext(basename)
            ext = ext.lower()
            if ext not in _ZIP_VALID_EXTENSIONS:
                skipped.append(basename)
                continue

            safe_name = os.path.basename(basename)

            if persist_dir:
                # Persist to session files/ directory so the file survives
                # page reloads and server restarts — applies to ALL file
                # types (.bm2, .csv, .txt, .xlsx), not just .bm2.
                file_path = str(persist_dir / safe_name)
                with zf.open(member) as src, open(file_path, "wb") as dst:
                    shutil.copyfileobj(src, dst)
                tmp_dir = None  # no temp dir needed — file is in session dir
            else:
                # No DTXSID provided — extract to a temp directory (lost on restart)
                tmp_dir = tempfile.mkdtemp(prefix=f"{ext.lstrip('.')}_")
                file_path = os.path.join(tmp_dir, safe_name)
                with zf.open(member) as src, open(file_path, "wb") as dst:
                    shutil.copyfileobj(src, dst)

            file_id = str(uuid.uuid4())

            if ext == ".bm2":
                # Register in _bm2_uploads — same as /api/upload-bm2.
                # .bm2 files are BMDExpress output (benchmark dose results),
                # which need the special _bm2_uploads store because
                # downstream endpoints (/api/process-bm2) reference them.
                _bm2_uploads[file_id] = {
                    "filename": safe_name,
                    "temp_path": file_path,
                    "table_data": None,
                    "bm2_json": None,     # populated on first preview/process
                    "narrative": None,
                }
                # Fingerprint for cross-validation
                v_issues = []
                if dtxsid:
                    fp = fingerprint_and_store(file_id, safe_name, file_path, "bm2", dtxsid)
                    v_issues = run_lightweight_validation(fp, dtxsid)
                bm2_results.append({
                    "id": file_id,
                    "filename": safe_name,
                    "validation_issues": v_issues,
                })

            else:
                # .csv, .txt, and .xlsx — raw dose-response experimental
                # data (animal IDs × endpoints, tab- or comma-separated)
                # or spreadsheet data.  These are *input* data suitable for
                # importing into BMDExpress, NOT gene-level BMD output.
                # We store them as generic file pool entries; the client
                # tracks them in uploadedFiles and the user chooses which
                # file to use for each report section.
                file_type = ext.lstrip(".")  # "csv", "txt", or "xlsx"

                # Store internally so downstream endpoints can access
                # the extracted file by ID if needed (e.g., /api/preview).
                _data_uploads[file_id] = {
                    "filename": safe_name,
                    "temp_path": file_path,
                    "type": file_type,
                }

                # Fingerprint for cross-validation
                v_issues = []
                if dtxsid:
                    fp = fingerprint_and_store(file_id, safe_name, file_path, file_type, dtxsid)
                    v_issues = run_lightweight_validation(fp, dtxsid)

                # Only send safe metadata to the client (no server paths)
                other_results.append({
                    "id": file_id,
                    "filename": safe_name,
                    "type": file_type,
                    "validation_issues": v_issues,
                })

    # Clean up the zip temp file (individual extracted files are in their
    # own temp dirs and persist until the server shuts down)
    shutil.rmtree(zip_tmp_dir, ignore_errors=True)

    # Collect all validation issues from individual file fingerprints
    # into a flat list for the response — the client shows these as
    # immediate feedback after zip extraction.
    all_validation_issues = []
    for r in bm2_results + other_results:
        all_validation_issues.extend(r.get("validation_issues", []))

    return JSONResponse({
        "bm2_files": bm2_results,
        "other_files": other_results,
        "skipped": skipped,
        "validation_issues": all_validation_issues,
    })


# ---------------------------------------------------------------------------
# POST /api/process-genomics — run gene set and gene BMD ranking
# ---------------------------------------------------------------------------

@router.post("/api/process-genomics")
async def api_process_genomics(request: Request):
    """
    Process an uploaded gene-level BMD CSV to produce the data for the
    Gene Set BMD Analysis and Gene BMD Analysis report sections.

    Takes a previously-uploaded CSV (via csv_id from /api/upload-csv)
    and runs:
      1. NIEHS quality filtering (fold change, goodness-of-fit, BMDU/BMDL)
      2. GO Biological Process gene set ranking by median BMD
      3. Individual gene ranking by BMD

    Input JSON:
      {
        "csv_id": "uuid-from-upload",
        "organ": "liver",
        "sex": "male",
        "compound_name": "PFHxSAm",
        "dose_unit": "mg/kg"
      }

    Returns JSON:
      {
        "gene_sets": [{rank, go_id, go_term, bmd_median, bmdl_median,
                       n_genes, genes, direction}, ...],
        "top_genes": [{rank, gene_symbol, full_name, bmd, bmdl, bmdu,
                       direction, fold_change}, ...],
        "total_responsive_genes": N,
        "organ": "liver",
        "sex": "male"
      }
    """
    _csv_uploads = get_csv_uploads()

    body = await request.json()
    csv_id = body.get("csv_id", "")
    organ = body.get("organ", "")
    sex = body.get("sex", "")

    if not csv_id:
        return JSONResponse(
            {"error": "csv_id is required"},
            status_code=400,
        )

    csv_upload = _csv_uploads.get(csv_id)
    if not csv_upload:
        return JSONResponse(
            {"error": f"Unknown csv_id: {csv_id}"},
            status_code=404,
        )

    df = csv_upload.get("df")
    if df is None or df.empty:
        return JSONResponse(
            {"error": "No valid data in the uploaded CSV"},
            status_code=400,
        )

    try:
        # Run gene set and gene ranking in a thread pool because
        # ToxKBQuerier makes DB queries that could be slow for large datasets
        loop = asyncio.get_running_loop()

        def _run_genomics():
            # Open a read-only connection to the knowledge base for GO term
            # lookups.  The KB is the bmdx.duckdb file in the project root.
            db_path = str(Path(__file__).parent / "bmdx.duckdb")
            kb = ToxKBQuerier(db_path)
            try:
                gene_sets = rank_go_sets_by_bmd(df, kb, top_n=10)
                top_genes = rank_genes_by_bmd(df, top_n=10)

                # Fetch descriptions for the NIEHS dense 9pt blocks.
                # GO definitions from EBI QuickGO, gene summaries from
                # MyGene.info (human orthologs for full descriptions).
                go_ids = [gs["go_id"] for gs in gene_sets]
                go_descs = fetch_go_descriptions(go_ids, kb=kb)

                gene_syms = [g["gene_symbol"] for g in top_genes]
                gene_descs = fetch_gene_descriptions(gene_syms)

                return gene_sets, top_genes, go_descs, gene_descs
            finally:
                kb.close()

        gene_sets, top_genes, go_descs, gene_descs = await loop.run_in_executor(
            None, _run_genomics,
        )

        return JSONResponse({
            "gene_sets": gene_sets,
            "top_genes": top_genes,
            "go_descriptions": go_descs,
            "gene_descriptions": gene_descs,
            "total_responsive_genes": len(df),
            "organ": organ,
            "sex": sex,
        })

    except Exception as e:
        return JSONResponse(
            {"error": f"Genomics processing failed: {e}"},
            status_code=500,
        )


# ---------------------------------------------------------------------------
# GET /api/preview/{file_id} — preview uploaded file contents
# ---------------------------------------------------------------------------
# Returns a JSON payload shaped for the browser's file-preview modal.
# The shape varies by file type so the front end can switch on `type`:
#   "bm2_json"  — processed .bm2: includes tables_json + narrative
#   "bm2_raw"   — unprocessed .bm2: just a "not yet processed" message
#   "table"     — .csv or .txt: first 50 rows parsed into headers + rows
#   "info"      — .xlsx or unknown: filename and size only

@router.get("/api/preview/{file_id}")
async def api_preview_file(file_id: str):
    """
    Return preview-ready content for an uploaded file.

    Lookup order: check _bm2_uploads first (apical endpoint .bm2 files),
    then _data_uploads (raw dose-response .csv/.txt/.xlsx from zip archives).
    Returns 404 if the file_id is not found in either store.

    Response shape depends on file type — see module-level comment above.
    """
    _bm2_uploads = get_bm2_uploads()
    _data_uploads = get_data_uploads()
    _csv_uploads = get_csv_uploads()

    # --- Check .bm2 uploads first ---
    bm2_entry = _bm2_uploads.get(file_id)
    if bm2_entry:
        filename = bm2_entry["filename"]
        bm2_json = bm2_entry.get("bm2_json")

        if bm2_json is None:
            # In-memory cache is empty.  Try the LMDB B-tree cache first —
            # the bm2_cache module stores deserialized BMDProject dicts via
            # LMDB + orjson.  After first access, reads are memory-mapped
            # (OS page cache) + orjson.loads (~10ms for 5 MB).  This is the
            # fast path for previewing a file after page reload.
            bm2_path = bm2_entry["temp_path"]
            if not os.path.exists(bm2_path):
                return JSONResponse(
                    {"error": f"Uploaded file no longer exists: {filename}"},
                    status_code=410,
                )

            # Fast path — LMDB B-tree lookup.  Only loads the BMDProject
            # JSON (skips category analysis, stats, and narrative).
            loop = asyncio.get_running_loop()
            bm2_json = await loop.run_in_executor(
                None, bm2_cache.get_json, bm2_path,
            )
            if bm2_json is not None:
                bm2_entry["bm2_json"] = bm2_json

            if bm2_json is None:
                # Not in LMDB — run the full export-and-analyze pipeline.
                # This launches Java (slow) but also populates table_data
                # and narrative, so later process-bm2 calls are instant.
                # The pipeline stores the result in LMDB automatically.
                try:
                    table_data, _cat_lookup, bm2_json = await loop.run_in_executor(
                        None, build_table_data_from_bm2, bm2_path,
                    )
                    bm2_entry["table_data"] = table_data
                    bm2_entry["bm2_json"] = bm2_json

                    # Generate a default narrative (will be overwritten with
                    # real compound name / dose unit when the user processes
                    # via the section builder)
                    narrative = generate_results_narrative(
                        table_data, "Test Compound", "mg/kg",
                    )
                    bm2_entry["narrative"] = narrative
                except Exception as e:
                    return JSONResponse(
                        {"error": f"Processing failed: {e}"},
                        status_code=500,
                    )

        # Return the full BMDProject JSON — contains all 7 top-level lists:
        #   doseResponseExperiments, bMDResult, categoryAnalysisResults,
        #   oneWayANOVAResults, williamsTrendResults, curveFitPrefilterResults,
        #   oriogenResults.
        # The JSON tree viewer in the modal lets users expand/collapse any
        # section, so they can inspect BMD model fits, category analysis
        # results, etc. — not just the dose-response tables.
        # Use orjson directly instead of JSONResponse (stdlib json.dumps).
        # For a 5-10 MB BMDProject dict, orjson.dumps is ~10x faster
        # (~15ms vs ~200ms), making the second preview feel instant.
        payload = {"type": "bm2_json", "filename": filename, "data": bm2_json or {}}
        return Response(
            content=orjson.dumps(payload),
            media_type="application/json",
        )

    # --- Check data uploads (.csv, .txt, .xlsx from zip extraction) ---
    data_entry = _data_uploads.get(file_id)
    if data_entry:
        filename = data_entry["filename"]
        temp_path = data_entry["temp_path"]
        file_type = data_entry.get("type", "")

        # Tabular files: parse first 50 rows into headers + rows arrays
        if file_type in ("csv", "txt"):
            try:
                separator = "," if file_type == "csv" else "\t"
                with open(temp_path, "r", encoding="utf-8", errors="replace") as f:
                    all_lines = f.readlines()

                # Strip trailing newlines and skip completely empty lines
                all_lines = [ln.rstrip("\n\r") for ln in all_lines if ln.strip()]
                total_rows = max(len(all_lines) - 1, 0)  # minus header row

                headers = all_lines[0].split(separator) if all_lines else []
                # Return up to 50 data rows (excluding the header)
                rows = [
                    line.split(separator)
                    for line in all_lines[1:51]
                ]

                return JSONResponse({
                    "type": "table",
                    "filename": filename,
                    "headers": headers,
                    "rows": rows,
                    "total_rows": total_rows,
                })
            except Exception as e:
                logger.warning("Preview parse failed for %s: %s", filename, e)
                return JSONResponse({
                    "type": "info",
                    "filename": filename,
                    "error": f"Could not parse file: {e}",
                })

        # Excel files: parse each worksheet with openpyxl and return
        # headers + first 50 rows per sheet for tabular preview.
        if file_type == "xlsx":
            try:
                import openpyxl
                # Note: read_only=True is avoided here because some xlsx files
                # (e.g., NTP/NIEHS exports) lack dimension metadata, causing
                # openpyxl's read-only mode to see only 1 row per sheet.
                wb = openpyxl.load_workbook(temp_path, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows_iter = ws.iter_rows(values_only=True)
                    # First row is treated as column headers
                    first_row = next(rows_iter, None)
                    headers = [str(c) if c is not None else "" for c in first_row] if first_row else []
                    # Collect up to 50 data rows for the preview
                    data_rows = []
                    for i, row in enumerate(rows_iter):
                        if i >= 50:
                            break
                        data_rows.append([str(c) if c is not None else "" for c in row])
                    # Total row count (subtract 1 for the header row)
                    total = max(ws.max_row - 1, 0) if ws.max_row else 0
                    sheets.append({
                        "name": sheet_name,
                        "headers": headers,
                        "rows": data_rows,
                        "total_rows": total,
                    })
                wb.close()
                return JSONResponse({
                    "type": "xlsx_table",
                    "filename": filename,
                    "sheets": sheets,
                })
            except Exception as e:
                logger.warning("XLSX preview failed for %s: %s", filename, e)
                return JSONResponse({
                    "type": "info",
                    "filename": filename,
                    "error": f"Could not parse xlsx: {e}",
                })

        # Unknown type fallback
        return JSONResponse({
            "type": "info",
            "filename": filename,
            "message": f"Preview not available for .{file_type} files.",
        })

    # --- Also check _csv_uploads (gene-level BMD CSVs) ---
    csv_entry = _csv_uploads.get(file_id)
    if csv_entry:
        filename = csv_entry["filename"]
        temp_path = csv_entry["temp_path"]
        try:
            with open(temp_path, "r", encoding="utf-8", errors="replace") as f:
                all_lines = f.readlines()
            all_lines = [ln.rstrip("\n\r") for ln in all_lines if ln.strip()]
            total_rows = max(len(all_lines) - 1, 0)
            headers = all_lines[0].split(",") if all_lines else []
            rows = [line.split(",") for line in all_lines[1:51]]
            return JSONResponse({
                "type": "table",
                "filename": filename,
                "headers": headers,
                "rows": rows,
                "total_rows": total_rows,
            })
        except Exception as e:
            return JSONResponse({
                "type": "info",
                "filename": filename,
                "error": f"Could not parse file: {e}",
            })

    # Not found in any store
    return JSONResponse(
        {"error": f"File not found: {file_id}"},
        status_code=404,
    )
