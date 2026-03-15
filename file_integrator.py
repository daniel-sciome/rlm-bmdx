"""
file_integrator — File fingerprinting and cross-validation for pool files.

Every file uploaded to a DTXSID session's pool represents experimental data
at one of three processing tiers:

    xlsx (NTP study team raw)  →  txt/csv (BMDExpress-importable pivot)  →  .bm2 (BMDExpress output)

These tiers describe the *same underlying experimental data* at different
processing stages.  This module fingerprints each file on addition and
cross-validates the pool to detect:

  - Coverage gaps (xlsx exists for Tissue Concentration, but no .bm2 analysis)
  - Structural contradictions (different dose groups or animal counts between tiers)
  - Redundant files (male_body_weight.csv and male_body_weight.txt identical)

Data precedence (default, user-overridable):
    xlsx files from the study team are authoritative.  The chain is:
      1. xlsx  — ground truth (direct from study team)
      2. txt/csv — derived from xlsx (reformatted for BMDExpress import)
      3. bm2   — derived from txt/csv (BMDExpress analysis output)

Primary identifiers:
  - DTXSID is the primary test article identifier
  - IUPAC name is the source of truth for chemical structure

Usage:
    from file_integrator import fingerprint_file, validate_pool, lightweight_validate

    # Fingerprint a single file
    fp = fingerprint_file(file_id, filename, path, "xlsx", ts_added)

    # Full pool validation
    report = validate_pool(dtxsid, fingerprints)

    # Lightweight check on file addition
    issues = lightweight_validate(new_fingerprint, existing_fingerprints)
"""

# ---------------------------------------------------------------------------
# Imports
# ---------------------------------------------------------------------------

import json
import logging
import os
import re
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

logger = logging.getLogger(__name__)

# Tier numbers encode data precedence: lower = more authoritative.
# xlsx is direct from study teams, txt/csv are reformatted for BMDExpress,
# bm2 is BMDExpress output (most derived, least authoritative by default).
TIER_XLSX = 1
TIER_TXT_CSV = 2
TIER_BM2 = 3

# Platform detection patterns — maps filename substrings (case-insensitive)
# to (platform, data_type) tuples.  The order matters: more specific patterns
# are checked first so "tissue_conc" doesn't accidentally match "clinical".
#
# platform uses the Apical platform vocabulary (e.g., "Body Weight",
# "Clinical Chemistry").  data_type is one of:
#   "tox_study"       — source-of-truth data with potential gaps
#   "inferred"        — gap-filled data for BMDExpress modeling
#   "gene_expression" — transcriptomic microarray/TempO-Seq data
#
# NTP naming conventions use both snake_case (in txt/csv filenames) and
# Title_Case (in xlsx filenames).  Each pattern is a tuple of
# (compiled_regex, platform, data_type) checked against the filename.
_PLATFORM_PATTERNS: list[tuple[re.Pattern, str | None, str]] = [
    # Gene expression — matches both "Gene_Expression" (bm2) and
    # organ-prefixed transcriptomics data like "Liver_PFHxSAm_Male_No0.txt".
    # The organ-prefixed pattern must be VERY large (>100KB) to avoid
    # false positives from small organ weight files.
    # platform is None because gene expression platform comes from chip info.
    (re.compile(r"gene.?expression", re.IGNORECASE), None, "gene_expression"),

    # --- Tox study files (data_type="tox_study") ---
    # Files with "_tox_study" (or legacy "_truth") in the name are source-of-
    # truth data: actual experimental values with potential gaps (missing
    # animals, lost samples).  Used for NTP traditional statistics (Williams,
    # Dunnett, Jonckheere) and for computing N, mean, SD per dose group.
    # These must come BEFORE the generic patterns so that
    # "body_weight_tox_study" matches here, not as "body_weight" below.
    # The _truth alias handles existing files that predate the rename.
    (re.compile(r"tissue.?conc.*_(?:tox_study|truth)", re.IGNORECASE), "Tissue Concentration", "tox_study"),
    (re.compile(r"clinical.?obs.*_(?:tox_study|truth)", re.IGNORECASE), "Clinical", "tox_study"),
    (re.compile(r"clin.*chem.*_(?:tox_study|truth)", re.IGNORECASE), "Clinical Chemistry", "tox_study"),
    (re.compile(r"hematol.*_(?:tox_study|truth)", re.IGNORECASE), "Hematology", "tox_study"),
    (re.compile(r"hormone.*_(?:tox_study|truth)", re.IGNORECASE), "Hormones", "tox_study"),
    (re.compile(r"organ.?weight.*_(?:tox_study|truth)", re.IGNORECASE), "Organ Weight", "tox_study"),
    (re.compile(r"body.?weight.*_(?:tox_study|truth)", re.IGNORECASE), "Body Weight", "tox_study"),

    # --- Inferred files (data_type="inferred") ---
    # Files WITHOUT "_tox_study" in the name are inferred data: gaps filled
    # with dose-group averages so BMDExpress can model them.  The .bm2
    # output from these becomes the source of BMD/BMDL values.
    # Tissue concentration — check before generic "clinical" to avoid
    # "tissue_conc" matching a broader "clin" pattern.
    (re.compile(r"tissue.?conc", re.IGNORECASE), "Tissue Concentration", "inferred"),

    # Clinical observations — "clinical_obs" or "Clinical_Observations"
    (re.compile(r"clinical.?obs", re.IGNORECASE), "Clinical", "inferred"),

    # Clinical chemistry — "clin_chem" or "Clinical_Chemistry" or "Clinical Chemistry"
    # The .* bridge handles both short forms ("clin_chem") and full words
    # ("Clinical_Chemistry") where "ical_" separates "Clin" and "Chem".
    (re.compile(r"clin.*chem", re.IGNORECASE), "Clinical Chemistry", "inferred"),

    # Hematology — "hematol" or "Hematology"
    (re.compile(r"hematol", re.IGNORECASE), "Hematology", "inferred"),

    # Hormones — "hormone" or "Hormone"
    (re.compile(r"hormone", re.IGNORECASE), "Hormones", "inferred"),

    # Organ weights — "organ_weight" or "Organ_Weight"
    (re.compile(r"organ.?weight", re.IGNORECASE), "Organ Weight", "inferred"),

    # Body weight — "body_weight" or "Body_Weight" or "Body weight"
    (re.compile(r"body.?weight", re.IGNORECASE), "Body Weight", "inferred"),
]

# ---------------------------------------------------------------------------
# Backward-compatibility helpers
# ---------------------------------------------------------------------------
# The old 14-domain model used suffixed names like "body_weight_tox_study"
# and "body_weight_inferred".  base_domain() stripped the suffix.
# Now that platform + data_type are separate fields, base_domain() is a
# no-op identity function kept only because it's imported by:
#   methods_report.py, animal_report.py, pool_integrator.py, pool_orchestrator.py
# Those callers will be updated separately.

_DOMAIN_SUFFIX_RE = re.compile(r"(_tox_study|_inferred)$")


def base_domain(domain: str) -> str:
    """
    DEPRECATED — returns its input unchanged.

    Previously stripped _tox_study / _inferred suffixes from the monolithic
    domain string.  Now that platform and data_type are separate fields on
    FileFingerprint, this function is a no-op.  Kept for backward
    compatibility with callers in other modules.

    Args:
        domain: Any domain/platform string.

    Returns:
        The same string, unchanged.
    """
    if not domain:
        return domain
    # Still strip for backward compat — callers pass old-style domain strings
    # until they're migrated to use platform directly.
    return _DOMAIN_SUFFIX_RE.sub("", domain)


# Gene expression txt files are organ-prefixed (e.g., "Liver_PFHxSAm_Male_No0.txt").
# They're distinguished from other small txt files by having >500 data rows
# (probes/genes) and organ-prefixed filenames.
_GENE_EXPR_ORGAN_PATTERN = re.compile(
    r"^(Liver|Kidney|Lung|Heart|Brain|Thymus|Spleen|Adrenal|Testis|Ovary|Thyroid)",
    re.IGNORECASE,
)

# Sex detection patterns for txt/csv filenames (e.g., "male_body_weight.txt").
# Uses word boundary (\b) but also matches at start-of-string — filenames
# like "male_body_weight.txt" start with "male" directly, and \b treats
# the underscore as a word character.  So we also match at string boundaries.
_SEX_PATTERN = re.compile(r"(?:^|[\b_\-\s])(male|female)(?:[\b_\-\s\.]|$)", re.IGNORECASE)

# BM2 experiment name patterns for sex detection (e.g., "BodyWeightMale")
_BM2_SEX_PATTERN = re.compile(r"(Male|Female)", re.IGNORECASE)

# BM2 experiment name → platform mapping.  These are the experiment name
# prefixes used by BMDExpress 3 when exporting apical endpoint data.
# Includes both full BMDExpress names (e.g., "ClinicalChemistry") and
# abbreviated forms from xlsx_to_pivot_txt output (e.g., "clin_chem" →
# stripped to "clinchem").  Both forms are needed because _partition_by_domain
# and _filter_gene_expression strip underscores before matching.
#
# Values are Apical platform vocabulary strings (e.g., "Body Weight").
# The data_type for all bm2 apical data is "inferred" (bm2 = BMDExpress
# output from gap-filled data).
_BM2_PLATFORM_MAP: dict[str, str] = {
    "bodyweight": "Body Weight",
    "organweight": "Organ Weight",
    "clinicalchemistry": "Clinical Chemistry",
    "clinchem": "Clinical Chemistry",
    "hematology": "Hematology",
    "hormone": "Hormones",
    "tissueconcentration": "Tissue Concentration",
    "tissueconc": "Tissue Concentration",
    "clinicalobservation": "Clinical",
    "clinicalobs": "Clinical",
    "clinobs": "Clinical",
}

# Backward-compatibility alias — imported by pool_orchestrator.py and
# pool_integrator.py.  Maps to the old-style domain strings so those
# modules keep working until they're migrated.
_BM2_DOMAIN_MAP: dict[str, str] = {
    # .bm2 files are always produced from inferred (gap-filled) data,
    # so they map to _inferred domains.  The source-of-truth data
    # (with gaps) arrives as separate txt/csv files.
    "bodyweight": "body_weight_inferred",
    "organweight": "organ_weights_inferred",
    "clinicalchemistry": "clin_chem_inferred",
    "clinchem": "clin_chem_inferred",
    "hematology": "hematology_inferred",
    "hormone": "hormones_inferred",
    "tissueconcentration": "tissue_conc_inferred",
    "tissueconc": "tissue_conc_inferred",
    # Clinical observations are categorical, not BMD-modeled — no _inferred variant
    "clinicalobservation": "clinical_obs",
    "clinicalobs": "clinical_obs",
    "clinobs": "clinical_obs",
}


# ---------------------------------------------------------------------------
# LLM-based metadata deduction for .bm2 experiment names
# ---------------------------------------------------------------------------
# The BMDExpress data model stores clinical endpoints identically to gene
# expression probes — "Alanine aminotransferase" is a "probe", "SD0" (study
# day 0) is a "probe".  Experiment names like "female_clin_chem" or
# "Kidney_PFHxSAm_Male_No0" encode metadata that regex alone can't reliably
# extract (especially test article names and strain information).
#
# We use a cheap, fast LLM call (Haiku) to parse experiment names into
# structured metadata: sex, organ, species, strain, domain, and test article.
# This mirrors BMDExpress-3's own LlmMetadataService (Java-side) and uses
# the same controlled vocabulary from vocabulary.yml.

# Controlled vocabulary — mirrors BMDExpress-3's vocabulary.yml so the LLM
# targets canonical values that match the augmented ExperimentDescription schema.
_VOCAB = {
    "species": [
        "rat", "mouse", "human", "rabbit", "dog",
        "monkey", "zebrafish", "guinea pig", "hamster", "pig",
    ],
    "strains": {
        "rat": ["Sprague-Dawley", "Wistar", "Long-Evans", "Fischer 344", "Brown Norway"],
        "mouse": ["C57BL/6", "BALB/c", "CD-1", "FVB/N", "129", "DBA/2", "NOD", "SCID"],
    },
    "sexes": ["male", "female", "both", "mixed", "NA"],
    "organs": [
        "adrenal", "blood", "bone", "brain", "colon", "heart", "intestine",
        "kidney", "liver", "lung", "muscle", "ovary", "pancreas", "prostate",
        "skin", "spleen", "stomach", "testes", "thymus", "thyroid", "uterus",
    ],
    # Platforms use the Apical platform vocabulary — these are the canonical
    # display names for experimental data categories.  data_type (tox_study,
    # inferred, gene_expression) is a separate concept not in this vocab.
    "platforms": [
        "Body Weight", "Organ Weight", "Clinical Chemistry",
        "Hematology", "Hormones", "Tissue Concentration",
        "Clinical", "gene_expression",
    ],
}

# Cache for LLM deduction results — keyed by experiment name string.
# Avoids redundant API calls when the same experiment names appear in
# multiple validation runs.  Persists for the process lifetime only.
_llm_cache: dict[str, dict] = {}


def _deduce_metadata_from_experiments(
    experiment_names: list[str],
    probe_ids_sample: list[str] | None = None,
    chip_info: dict | None = None,
) -> dict:
    """
    Use a fast LLM (Haiku) to extract structured metadata from experiment names.

    Given experiment names like ["female_clin_chem", "male_clin_chem"] or
    ["Kidney_PFHxSAm_Female_No0", "Liver_PFHxSAm_Male_No0"], the LLM
    returns structured JSON with:
      - sexes: list of sexes found across all experiments
      - organs: list of organs (for gene expression)
      - species: inferred species (from strain, chip, or naming conventions)
      - strain: inferred strain
      - platform: the Apical platform name or "gene_expression"
      - test_article: the chemical compound name if embedded in experiment names

    Falls back to regex-based detection if the LLM is unavailable or fails.

    Args:
        experiment_names:  List of experiment.name values from BMDProject JSON.
        probe_ids_sample:  Optional sample of probe/endpoint IDs (first 10) to
                           help the LLM distinguish clinical endpoints from genes.
        chip_info:         Optional chip metadata dict with provider, species, name.

    Returns:
        Dict with keys: sexes, organs, species, strain, platform, test_article.
        All values are strings or lists of strings; None if not detected.
    """
    if not experiment_names:
        return {}

    # Build a cache key from the experiment names
    cache_key = "|".join(sorted(experiment_names))
    if cache_key in _llm_cache:
        return _llm_cache[cache_key]

    # Build the prompt — modeled after BMDExpress-3's LlmMetadataService
    prompt = (
        "You extract experimental metadata from BMDExpress experiment names.\n"
        "Given the experiment names and optional context below, return ONLY a "
        "JSON object with these fields (use null if not determinable):\n\n"
        "  sexes       — list of sexes found (from vocabulary)\n"
        "  organs      — list of organs found (from vocabulary), or null\n"
        "  species     — single species string, or null\n"
        "  strain      — single strain string, or null\n"
        "  platform    — one of the platform values, or null\n"
        "  test_article — chemical/compound name if present, or null\n\n"
        "Valid values:\n"
        f"  species: {_VOCAB['species']}\n"
        f"  sexes: {_VOCAB['sexes']}\n"
        f"  organs: {_VOCAB['organs']}\n"
        f"  platforms: {_VOCAB['platforms']}\n"
        f"  strains by species: {_VOCAB['strains']}\n\n"
        "Rules:\n"
        "- Use vocabulary values above when there is a close match.\n"
        "- For test_article, extract the chemical/compound name; replace "
        "underscores with spaces.\n"
        "- If probe IDs are human-readable endpoint names (like 'Albumin', "
        "'Hematocrit', 'Heart'), this is a clinical/apical endpoint file, "
        "NOT gene expression.\n"
        "- If probe IDs look like gene symbols with numeric suffixes "
        "(like 'AADAC_7934', 'CYP1A1_12345'), this is gene expression.\n"
        "- 'SD0', 'SD5' are Study Day body weight measurements → platform = Body Weight.\n"
        "- Return ONLY valid JSON. No markdown, no explanation.\n\n"
        f"Experiment names: {experiment_names}\n"
    )

    if probe_ids_sample:
        prompt += f"Sample probe/endpoint IDs: {probe_ids_sample}\n"
    if chip_info and isinstance(chip_info, dict):
        # Include chip provider and species if available — helps distinguish
        # gene expression (BioSpyder, Affymetrix) from clinical endpoints.
        chip_summary = {
            k: v for k, v in chip_info.items()
            if k in ("provider", "species", "name") and v
        }
        if chip_summary:
            prompt += f"Chip/platform info: {chip_summary}\n"

    # Call Haiku via the Anthropic API — fast (~200ms) and cheap (~$0.0003)
    try:
        import anthropic
        client = anthropic.Anthropic()
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=512,
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip() if response.content else ""

        # Strip markdown code fences if present
        if raw.startswith("```"):
            first_nl = raw.index("\n") if "\n" in raw else 3
            raw = raw[first_nl + 1:]
            if raw.endswith("```"):
                raw = raw[:-3].strip()

        result = json.loads(raw)
        # Haiku sometimes returns a list of objects when there are multiple
        # experiments (e.g., gene expression with 4 organ×sex combos).
        # We expect a single dict — if we get a list, merge them.
        if isinstance(result, list):
            merged: dict = {}
            for item in result:
                if isinstance(item, dict):
                    for k, v in item.items():
                        if k not in merged or merged[k] is None:
                            merged[k] = v
                        elif isinstance(merged[k], list) and isinstance(v, list):
                            # Merge list values (e.g., sexes, organs)
                            merged[k] = sorted(set(merged[k] + v))
            result = merged
        if not isinstance(result, dict):
            logger.warning("LLM returned non-dict: %s", type(result))
            return {}
        _llm_cache[cache_key] = result
        logger.debug("LLM metadata for %s: %s", experiment_names[0], result)
        return result

    except Exception as e:
        logger.warning(
            "LLM metadata deduction failed for experiments %s: %s",
            experiment_names[:2], e,
        )
        # Fall back to empty dict — caller uses regex-based detection
        return {}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class FileFingerprint:
    """
    Extracted metadata from a single pool file, used for cross-validation.

    Every file in the pool — regardless of type (xlsx, txt, csv, bm2) —
    gets fingerprinted on addition.  The fingerprint captures enough
    structural information to detect overlaps and conflicts between files
    without re-reading the file contents.

    Fields:
      file_id       — UUID from upload (links back to _bm2_uploads/_data_uploads)
      filename      — original filename as uploaded
      file_type     — "xlsx", "txt", "csv", or "bm2"
      tier          — precedence rank: 1=xlsx, 2=txt/csv, 3=bm2

      ts_added      — ISO 8601 timestamp when file entered the pool
      ts_filesystem — ISO 8601 file mtime from the OS (None if unavailable)
      ts_internal   — ISO 8601 date found inside file content:
                        xlsx: from "Key to Column Labels" metadata (column D)
                        bm2:  from BMDProject creation date field (if present)
                        txt/csv: None (no internal dates)

      study_number  — e.g., "C20022-01" (from xlsx data or bm2 metadata)
      species       — e.g., "Sprague Dawley" (from bm2 experiment names)
      sexes         — ["Male", "Female"] or ["Male"] etc.
      dose_groups   — sorted unique doses, e.g., [0, 0.15, 0.5, 1.4, 4, 12, 37, 111]
      dose_unit     — "mg/kg" if detectable

      platform      — Apical platform name ("Body Weight", "Clinical Chemistry",
                        etc.) or None for gene expression / unrecognized files
      data_type     — "tox_study", "inferred", or "gene_expression"
      endpoint_names — e.g., ["SD0", "SD5"] or ["Liver Weight Absolute", ...]
      animal_ids    — e.g., ["101", "102", ...] — for cross-tier count comparison
      n_animals_by_dose — {0.0: 10, 4.0: 8, ...} — animals per dose group

      organ         — e.g., "Liver", "Kidney" (for gene expression files)
      gene_count    — number of probes/genes (for gene expression .txt and .bm2)
    """

    file_id: str
    filename: str
    file_type: str
    tier: int

    # Timestamps — for evaluating data truth precedence
    ts_added: str
    ts_filesystem: str | None = None
    ts_internal: str | None = None

    # Study structural fingerprint
    study_number: str | None = None
    species: str | None = None
    sexes: list[str] = field(default_factory=list)
    dose_groups: list[float] = field(default_factory=list)
    dose_unit: str | None = None

    # Endpoint coverage — platform is the Apical vocabulary name (e.g.,
    # "Body Weight"), data_type distinguishes tox_study vs inferred vs
    # gene_expression.  Together they replace the old monolithic domain string.
    platform: str | None = None
    data_type: str | None = None
    endpoint_names: list[str] = field(default_factory=list)
    animal_ids: list[str] = field(default_factory=list)
    n_animals_by_dose: dict[float, int] = field(default_factory=dict)

    # Study-file detection — True when the xlsx has the two-tab NTP structure:
    # "Key to Column Labels" + "Data" sheets.  Study files are authoritative
    # for doses and animal roster; bm2 files are authoritative for BMD results.
    is_study_file: bool = False

    # Selection groups found in the xlsx "Selection" column (e.g., "Core Animals",
    # "Biosampling Animals").  Used to filter to core animals for the report.
    selection_groups: list[str] = field(default_factory=list)

    # Animals per dose group per selection group.  Outer key is dose, inner
    # maps selection label → set of animal IDs.  Enables detecting which
    # dose groups lost animals (dead before terminal sacrifice).
    animals_by_dose_selection: dict[float, dict[str, list[str]]] = field(default_factory=dict)

    # Core animals per dose × sex — the reference roster for dead-animal
    # detection.  Only populated for study xlsx files with a "Selection"
    # column.  Structure: {dose → {"Male": [aids], "Female": [aids]}}.
    core_animals_by_dose_sex: dict[float, dict[str, list[str]]] = field(default_factory=dict)

    # Long-format flag — True when a txt/csv file has NTP long format
    # (one row per animal × observation, with Concentration/Animal ID/Sex
    # columns).  These need conversion to wide format before integration.
    is_long_format: bool = False

    # Transcriptomics-specific
    organ: str | None = None
    gene_count: int | None = None


@dataclass
class ValidationIssue:
    """
    A single discrepancy or coverage gap found during cross-validation.

    severity levels:
      "error"   — structural conflict requiring user resolution
                  (e.g., dose groups differ between xlsx and txt)
      "warning" — coverage gap or potential issue
                  (e.g., no bm2 analysis for a domain that has txt data)
      "info"    — informational note, no action needed
                  (e.g., xlsx exists but not all domains have txt/csv)

    issue_type values:
      "dose_mismatch"         — dose_groups differ across tiers in the same domain
      "animal_count_mismatch" — n_animals_by_dose differ (animals added or removed)
      "missing_tier"          — a domain is missing one or more expected tiers
      "redundant_files"       — two files have identical structural fingerprints
      "endpoint_mismatch"     — endpoint names don't match across tiers
      "sex_mismatch"          — sex coverage differs between tiers
    """

    severity: str
    platform: str
    issue_type: str
    message: str
    files_involved: list[str]
    suggested_precedence: str | None = None
    details: dict = field(default_factory=dict)


@dataclass
class ValidationReport:
    """
    Full cross-validation report for a session's file pool.

    coverage_matrix maps domain → { "xlsx": file_id|None, "txt_csv": [file_ids], "bm2": file_id|None }.
    The txt_csv tier can have multiple files because apical data is split by sex
    (e.g., male_body_weight.txt + female_body_weight.txt).

    is_complete is True only if every domain that has any tier also has all
    applicable tiers.  Gene expression typically has no xlsx.
    """

    dtxsid: str
    run_at: str
    file_count: int
    fingerprints: dict[str, dict]         # file_id → fingerprint as dict
    issues: list[dict]                     # list of ValidationIssue as dicts
    coverage_matrix: dict[str, dict]       # domain → tier → file_id(s)
    is_complete: bool


# ---------------------------------------------------------------------------
# Domain detection
# ---------------------------------------------------------------------------


def detect_platform_and_type(
    filename: str,
    file_type: str,
    file_size: int = 0,
    endpoint_names: list[str] | None = None,
) -> tuple[str | None, str | None]:
    """
    Infer the platform and data_type from a filename using NTP naming conventions.

    Checks filename against a prioritized list of regex patterns.  More
    specific patterns (Tissue Concentration, Clinical Observations) are
    checked before generic ones (Body Weight) to avoid false matches.

    For large txt files (>100KB) with organ-prefixed names (e.g.,
    "Liver_PFHxSAm_Male_No0.txt"), returns (None, "gene_expression") —
    these are transcriptomic microarray data, not apical endpoint tables.

    Args:
        filename:       Original filename as uploaded.
        file_type:      "xlsx", "txt", "csv", or "bm2".
        file_size:      File size in bytes (used for gene expression heuristic).
        endpoint_names: Parsed endpoint/row labels (not used currently, reserved
                        for future content-based detection).

    Returns:
        Tuple of (platform, data_type).  platform is an Apical vocabulary
        string like "Body Weight" or None for gene expression / unrecognized.
        data_type is "tox_study", "inferred", or "gene_expression", or None
        if unrecognized.
    """
    # Check named patterns first — covers most NTP file naming conventions
    for pattern, platform, data_type in _PLATFORM_PATTERNS:
        if pattern.search(filename):
            return (platform, data_type)

    # Gene expression heuristic for organ-prefixed txt files.
    # NTP transcriptomics data is named like "Liver_PFHxSAm_Male_No0.txt"
    # and is typically 500KB-1MB (thousands of probes × dozens of animals).
    # We require >100KB to avoid matching small organ weight txt files.
    if file_type in ("txt", "csv") and file_size > 100_000:
        if _GENE_EXPR_ORGAN_PATTERN.match(filename):
            return (None, "gene_expression")

    return (None, None)


def detect_domain(filename: str, file_type: str, file_size: int = 0,
                  endpoint_names: list[str] | None = None) -> str | None:
    """
    DEPRECATED backward-compatibility wrapper around detect_platform_and_type().

    Returns a combined domain string in the old format (e.g.,
    "body_weight_inferred", "gene_expression") for callers that haven't
    been migrated yet.  Imported by: animal_report.py, pool_orchestrator.py.

    Args:
        filename:       Original filename as uploaded.
        file_type:      "xlsx", "txt", "csv", or "bm2".
        file_size:      File size in bytes (used for gene expression heuristic).
        endpoint_names: Parsed endpoint/row labels (reserved for future use).

    Returns:
        Old-style domain string or None if unrecognized.
    """
    platform, data_type = detect_platform_and_type(
        filename, file_type, file_size, endpoint_names,
    )
    return _platform_type_to_domain(platform, data_type)


def _platform_type_to_domain(
    platform: str | None,
    data_type: str | None,
) -> str | None:
    """
    Convert a (platform, data_type) pair back to the old monolithic domain string.

    This is a backward-compatibility helper used by detect_domain() and
    detect_domain_from_bm2() to produce old-style strings for callers that
    haven't been migrated.  Will be removed once all callers use platform
    + data_type directly.

    The mapping reverses _PLATFORM_PATTERNS:
        ("Body Weight", "tox_study")   → "body_weight_tox_study"
        ("Body Weight", "inferred")    → "body_weight_inferred"
        (None, "gene_expression")      → "gene_expression"
        ("Clinical", *)   → "clinical_obs"

    Args:
        platform:  Apical platform name or None.
        data_type: "tox_study", "inferred", or "gene_expression", or None.

    Returns:
        Old-style domain string, or None if both inputs are None.
    """
    if platform is None and data_type is None:
        return None
    if data_type == "gene_expression":
        return "gene_expression"

    # Map platform display names to old-style base domain slugs
    _PLATFORM_TO_SLUG: dict[str, str] = {
        "Body Weight": "body_weight",
        "Organ Weight": "organ_weights",
        "Clinical Chemistry": "clin_chem",
        "Hematology": "hematology",
        "Hormones": "hormones",
        "Tissue Concentration": "tissue_conc",
        "Clinical": "clinical_obs",
    }

    slug = _PLATFORM_TO_SLUG.get(platform or "", platform or "unknown")

    # Clinical observations has no _inferred/_tox_study variant in old model
    if slug == "clinical_obs":
        return "clinical_obs"
    if data_type:
        return f"{slug}_{data_type}"
    return slug


def detect_platform_and_type_from_bm2(
    experiment_names: list[str],
) -> tuple[str | None, str | None]:
    """
    Infer (platform, data_type) from BMDExpress experiment names.

    BMDExpress 3 names experiments like "BodyWeightMale", "OrganWeightFemale",
    "HematologyMale", etc.  We normalize to lowercase and strip sex suffixes
    to match against _BM2_PLATFORM_MAP.

    For gene expression .bm2 files, the experiment names reference individual
    probes/genes rather than named endpoints — these have thousands of
    experiments, which is the distinguishing heuristic.

    All bm2 apical data has data_type="inferred" because bm2 files are
    always produced from gap-filled data.

    Args:
        experiment_names: List of experiment.name values from the BMDProject JSON.

    Returns:
        Tuple of (platform, data_type).  Gene expression → (None, "gene_expression").
        Apical → (platform_name, "inferred").  Unrecognized → (None, None).
    """
    if not experiment_names:
        return (None, None)

    # Gene expression bm2 files have thousands of "experiments" (one per probe).
    # Apical bm2 files have at most ~50 (endpoints × sexes).
    if len(experiment_names) > 200:
        return (None, "gene_expression")

    # Normalize the first experiment name: lowercase, strip sex suffix
    name = experiment_names[0].lower()
    # Remove trailing "male" or "female"
    name = re.sub(r"(male|female)$", "", name).strip()
    # Remove underscores, spaces, hyphens for fuzzy matching
    name_normalized = re.sub(r"[_\s\-]", "", name)

    for prefix, platform in _BM2_PLATFORM_MAP.items():
        if name_normalized.startswith(prefix):
            return (platform, "inferred")

    return (None, None)


def detect_domain_from_bm2(experiment_names: list[str]) -> str | None:
    """
    DEPRECATED backward-compatibility wrapper around detect_platform_and_type_from_bm2().

    Returns an old-style domain string for callers that haven't been
    migrated.  Imported by pool_orchestrator.py (indirectly via detect_domain).

    Args:
        experiment_names: List of experiment.name values from the BMDProject JSON.

    Returns:
        Old-style domain string or None if unrecognized.
    """
    platform, data_type = detect_platform_and_type_from_bm2(experiment_names)
    return _platform_type_to_domain(platform, data_type)


# ---------------------------------------------------------------------------
# Sex detection helpers
# ---------------------------------------------------------------------------


def _detect_sex_from_filename(filename: str) -> list[str]:
    """
    Extract sex indicators from a filename.

    NTP txt/csv files encode sex in the filename: "male_body_weight.txt",
    "female_clin_chem.txt".  Returns the list of sexes found (["Male"],
    ["Female"], or [] if neither is detected).

    Args:
        filename: The original filename.

    Returns:
        List of sex strings with title case (e.g., ["Male"]).
    """
    matches = _SEX_PATTERN.findall(filename)
    return sorted(set(m.title() for m in matches))


def _detect_sexes_from_bm2_experiments(experiment_names: list[str]) -> list[str]:
    """
    Extract sex indicators from BMDExpress experiment names.

    Experiment names encode sex as a suffix: "BodyWeightMale",
    "HematologyFemale".  Returns the unique sexes found.

    Args:
        experiment_names: List of experiment.name values from BMDProject JSON.

    Returns:
        Sorted list of sex strings (e.g., ["Female", "Male"]).
    """
    sexes = set()
    for name in experiment_names:
        matches = _BM2_SEX_PATTERN.findall(name)
        for m in matches:
            sexes.add(m.title())
    return sorted(sexes)


# ---------------------------------------------------------------------------
# Timestamp helpers
# ---------------------------------------------------------------------------


def _get_filesystem_timestamp(path: str) -> str | None:
    """
    Get the file's modification time as an ISO 8601 string.

    Returns None if the file doesn't exist or stat fails.

    Args:
        path: Absolute path to the file on disk.

    Returns:
        ISO 8601 timestamp string or None.
    """
    try:
        mtime = os.path.getmtime(path)
        return datetime.fromtimestamp(mtime, tz=timezone.utc).isoformat()
    except OSError:
        return None


# ---------------------------------------------------------------------------
# Fingerprint extractors — one per file type
# ---------------------------------------------------------------------------


def fingerprint_xlsx(
    file_id: str,
    filename: str,
    path: str,
    ts_added: str,
) -> FileFingerprint:
    """
    Parse both sheets of an NTP xlsx to extract study metadata + animal/dose structure.

    NTP xlsx files have two sheets:
      1. "Key to Column Labels" — metadata describing each column's meaning.
         Row 1 has column labels in col A and human descriptions in col B.
         Column D sometimes has study-level metadata:
           Row 2: study type (e.g., "TOX")
           Row 3: study number (e.g., "C20022-01")
           Row 4: date (e.g., "12/08/2020")

      2. "Data" — individual animal measurements in long format.
         Each row is one animal × one observation.  Columns vary by domain:
           Col A: NTP Study Number (e.g., "C20022-01")
           Col B: Concentration (dose as string, e.g., "0", "4", "111")
           Col C: Animal ID (e.g., "101", "102")
           Col D: Sex ("Male" or "Female")
           Remaining columns: domain-specific (body weight, endpoints, etc.)

    Args:
        file_id:   UUID from upload.
        filename:  Original filename as uploaded.
        path:      Absolute path to the xlsx file on disk.
        ts_added:  ISO 8601 timestamp when the file entered the pool.

    Returns:
        FileFingerprint with extracted study metadata, doses, animals, endpoints.
    """
    import openpyxl

    fp = FileFingerprint(
        file_id=file_id,
        filename=filename,
        file_type="xlsx",
        tier=TIER_XLSX,
        ts_added=ts_added,
        ts_filesystem=_get_filesystem_timestamp(path),
    )

    file_size = os.path.getsize(path) if os.path.exists(path) else 0
    fp.platform, fp.data_type = detect_platform_and_type(filename, "xlsx", file_size)

    try:
        # read_only=False because NTP xlsx files sometimes lack dimension
        # metadata, causing read-only mode to see only 1 row per sheet.
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        logger.warning("Could not open xlsx %s: %s", filename, e)
        return fp

    # --- Detect study file by sheet structure ---
    # NTP study xlsx files always have exactly two tabs: "Key to Column Labels"
    # and "Data".  This is more reliable than filename patterns for identifying
    # the authoritative study data source.
    has_key_sheet = any("Key to" in s for s in wb.sheetnames)
    has_data_sheet = "Data" in wb.sheetnames
    fp.is_study_file = has_key_sheet and has_data_sheet

    # --- Parse "Key to Column Labels" sheet for study metadata ---
    # Column D of this sheet sometimes has: study type (row 2), study number
    # (row 3), and date (row 4).
    try:
        key_sheet = wb[wb.sheetnames[0]]
        for row in key_sheet.iter_rows(values_only=True):
            vals = [v for v in row if v is not None]
            if not vals:
                continue
            # Column D (index 3) in key sheet has study-level metadata
            if len(row) >= 4 and row[3] is not None:
                val = str(row[3]).strip()
                # Study number pattern: letter + digits + hyphen + digits
                if re.match(r"^[A-Z]\d+-\d+$", val):
                    fp.study_number = val
                # Date pattern: MM/DD/YYYY
                elif re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", val):
                    try:
                        dt = datetime.strptime(val, "%m/%d/%Y")
                        fp.ts_internal = dt.isoformat()
                    except ValueError:
                        pass
    except Exception as e:
        logger.debug("Key sheet parse failed for %s: %s", filename, e)

    # --- Parse "Data" sheet for animal/dose structure ---
    # Columns: A=StudyNumber, B=Concentration, C=AnimalID, D=Sex, E+=endpoints
    try:
        data_sheet = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[-1]]
        headers = None
        doses_set: set[float] = set()
        sexes_set: set[str] = set()
        animal_ids_set: set[str] = set()
        # Track animal IDs per dose for n_animals_by_dose
        animals_per_dose: dict[float, set[str]] = {}
        # Track animals per dose × selection group (e.g., "Core Animals"
        # vs "Biosampling Animals").  Enables filtering to core animals
        # and detecting which dose groups lost animals.
        animals_by_dose_sel: dict[float, dict[str, set[str]]] = {}
        selections_set: set[str] = set()
        endpoint_names_set: set[str] = set()
        selection_col_idx: int | None = None
        # Core animals by dose × sex for dead-animal detection
        core_by_dose_sex: dict[float, dict[str, set[str]]] = {}

        for row_idx, row in enumerate(data_sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                # First row is column headers — capture endpoint column names.
                # Standard NTP columns are StudyNumber, Concentration, AnimalID,
                # Sex, Selection; anything after that is domain-specific.
                headers = [str(c) if c is not None else "" for c in row]
                # Find the Selection column index (usually col 4 / index 4,
                # but may shift if the file has extra metadata columns).
                for ci, h in enumerate(headers):
                    if h.strip().lower() == "selection":
                        selection_col_idx = ci
                        break
                if len(headers) > 5:
                    endpoint_names_set.update(headers[5:])
                elif len(headers) > 4:
                    endpoint_names_set.update(headers[4:])
                continue

            # Data rows — extract dose, animal ID, sex, and selection group
            if row[1] is not None:
                try:
                    dose = float(row[1])
                    doses_set.add(dose)
                    # Track animal IDs per dose
                    aid = str(row[2]) if row[2] is not None else ""
                    if aid:
                        if dose not in animals_per_dose:
                            animals_per_dose[dose] = set()
                        animals_per_dose[dose].add(aid)
                        # Track by selection group
                        sel_label = ""
                        if selection_col_idx is not None and len(row) > selection_col_idx:
                            sel_label = str(row[selection_col_idx]).strip() if row[selection_col_idx] else ""
                        if sel_label:
                            selections_set.add(sel_label)
                            animals_by_dose_sel.setdefault(dose, {}).setdefault(sel_label, set()).add(aid)
                        # Track Core Animals by dose × sex for dead-animal detection.
                        # "Core Animals" is the NTP designation for animals intended
                        # for terminal sacrifice (vs "Biosampling Animals" sacrificed
                        # at an interim timepoint).  If no Selection column exists,
                        # treat all animals as core.
                        sex_val = str(row[3]).strip().title() if row[3] else ""
                        is_core = (sel_label == "Core Animals") or (selection_col_idx is None)
                        if is_core and sex_val:
                            core_by_dose_sex.setdefault(dose, {}).setdefault(sex_val, set()).add(aid)
                except (ValueError, TypeError):
                    pass
            if row[2] is not None:
                animal_ids_set.add(str(row[2]))
            if row[3] is not None:
                sexes_set.add(str(row[3]).strip().title())
            # Study number from data rows (fallback if key sheet didn't have it)
            if fp.study_number is None and row[0] is not None:
                val = str(row[0]).strip()
                if re.match(r"^[A-Z]\d+-\d+$", val):
                    fp.study_number = val

        fp.dose_groups = sorted(doses_set)
        fp.sexes = sorted(sexes_set)
        fp.animal_ids = sorted(animal_ids_set)
        # Deduplicate animal counts — xlsx has multiple rows per animal (one per
        # observation day), so we count *unique* animal IDs per dose.
        fp.n_animals_by_dose = {
            dose: len(aids) for dose, aids in sorted(animals_per_dose.items())
        }
        # Selection groups and per-dose animal rosters
        fp.selection_groups = sorted(selections_set)
        fp.animals_by_dose_selection = {
            dose: {sel: sorted(aids) for sel, aids in sel_map.items()}
            for dose, sel_map in sorted(animals_by_dose_sel.items())
        }
        # Core animals by dose × sex — the authoritative roster for
        # dead-animal footnotes.
        fp.core_animals_by_dose_sex = {
            dose: {sex: sorted(aids) for sex, aids in sex_map.items()}
            for dose, sex_map in sorted(core_by_dose_sex.items())
        }
        # Endpoint names — filter out standard NTP columns
        standard_cols = {
            "NTP Study Number", "Concentration", "Animal ID", "Sex",
            "Selection", "Terminal Flag", "",
        }
        fp.endpoint_names = sorted(
            endpoint_names_set - standard_cols
        )
        # Dose unit — NTP xlsx "Key to Column Labels" always says "mg/kg"
        if any("mg/kg" in str(row) for row in key_sheet.iter_rows(values_only=True)):
            fp.dose_unit = "mg/kg"

    except Exception as e:
        logger.debug("Data sheet parse failed for %s: %s", filename, e)

    wb.close()
    return fp


def fingerprint_txt_csv(
    file_id: str,
    filename: str,
    path: str,
    file_type: str,
    ts_added: str,
) -> FileFingerprint:
    """
    Parse the BMDExpress-importable pivot-table format.

    NTP txt/csv files are transposed tables where:
      Row 1: Animal IDs (tab- or comma-separated), first cell blank or label
      Row 2: Dose concentrations per animal (same column order as row 1)
      Row 3+: Endpoint measurements — first cell is the endpoint name,
              remaining cells are values per animal

    For gene expression files (Liver_PFHxSAm_Male_No0.txt), row 1 has
    "Something" then plate-prefixed sample IDs, row 2 has "Doses" then
    concentrations, and rows 3+ are probe/gene expression values.

    Args:
        file_id:    UUID from upload.
        filename:   Original filename as uploaded.
        path:       Absolute path to the txt/csv file on disk.
        file_type:  "txt" or "csv" (determines separator character).
        ts_added:   ISO 8601 timestamp when the file entered the pool.

    Returns:
        FileFingerprint with doses, animal IDs, endpoints, sex, and domain.
    """
    fp = FileFingerprint(
        file_id=file_id,
        filename=filename,
        file_type=file_type,
        tier=TIER_TXT_CSV,
        ts_added=ts_added,
        ts_filesystem=_get_filesystem_timestamp(path),
    )

    # Detect sex from filename — NTP convention: "male_body_weight.txt"
    fp.sexes = _detect_sex_from_filename(filename)

    file_size = os.path.getsize(path) if os.path.exists(path) else 0
    fp.platform, fp.data_type = detect_platform_and_type(filename, file_type, file_size)

    separator = "," if file_type == "csv" else "\t"

    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
    except Exception as e:
        logger.warning("Could not read %s: %s", filename, e)
        return fp

    # Strip trailing whitespace and skip empty lines
    lines = [ln.rstrip("\n\r") for ln in lines if ln.strip()]
    if len(lines) < 2:
        return fp

    # ----- Detect long format (NTP study data) vs wide format (BMDExpress) -----
    # Long format: header row has named columns like "NTP Study Number",
    # "Concentration", "Animal ID", "Sex".  Each data row is one observation.
    # Wide format: row 1 = animal IDs, row 2 = doses, rows 3+ = endpoints.
    row1_cells = lines[0].split(separator)
    row1_lower = {c.strip().lower() for c in row1_cells}

    # Detect long format by checking for characteristic column names
    _LONG_FORMAT_MARKERS = {"concentration", "animal id", "sex"}
    is_long_format = _LONG_FORMAT_MARKERS.issubset(row1_lower)

    if is_long_format:
        fp.is_long_format = True
        # --- Long-format parsing (source-of-truth _tox_study.csv files) ---
        # Parse header → column indices, then extract doses, animal IDs,
        # sexes, and endpoint names from the data rows.
        header = [c.strip() for c in row1_cells]
        header_lower = [h.lower() for h in header]

        # Find key column indices
        dose_col = header_lower.index("concentration")
        animal_col = header_lower.index("animal id")
        sex_col = header_lower.index("sex") if "sex" in header_lower else None

        # Endpoint columns: everything after the metadata columns.
        # Known metadata columns that are NOT endpoints.
        _META_COLS = {
            "ntp study number", "concentration", "animal id", "sex",
            "selection", "observation day", "terminal flag",
            # Clinical observations metadata
            "observation", "site", "modifier", "severity",
        }
        endpoint_cols = [
            i for i, h in enumerate(header_lower)
            if h not in _META_COLS and h
        ]

        doses_set: set[float] = set()
        animal_ids_set: set[str] = set()
        sexes_set: set[str] = set()
        dose_animal_map: dict[float, set[str]] = {}
        endpoint_names_set: set[str] = set()

        for line in lines[1:]:
            cells = line.split(separator)
            if len(cells) <= max(dose_col, animal_col):
                continue

            # Dose
            try:
                dose_val = float(cells[dose_col].strip())
                doses_set.add(dose_val)
                aid = cells[animal_col].strip()
                if aid:
                    animal_ids_set.add(aid)
                    dose_animal_map.setdefault(dose_val, set()).add(aid)
            except (ValueError, TypeError, IndexError):
                continue

            # Sex
            if sex_col is not None and sex_col < len(cells):
                s = cells[sex_col].strip()
                if s:
                    sexes_set.add(s.title())

        fp.dose_groups = sorted(doses_set)
        fp.animal_ids = sorted(animal_ids_set)
        fp.n_animals_by_dose = {
            dose: len(aids) for dose, aids in sorted(dose_animal_map.items())
        }

        # Override sex detection with actual data (more reliable than filename)
        if sexes_set:
            fp.sexes = sorted(sexes_set)

        # Endpoint names from the non-metadata columns
        fp.endpoint_names = [header[i] for i in endpoint_cols]

        return fp

    # ----- Wide-format parsing (BMDExpress input files) -----
    # Row 1: animal IDs (first cell is blank or a label like "Something")
    # Skip the first cell (it's a label column), rest are animal IDs
    animal_ids = [c.strip() for c in row1_cells[1:] if c.strip()]
    fp.animal_ids = sorted(set(animal_ids))

    # Row 2: dose concentrations per animal
    row2_cells = lines[1].split(separator)
    doses_raw = row2_cells[1:]  # skip label cell
    doses_set: set[float] = set()
    # Map dose → set of animal IDs for n_animals_by_dose
    dose_animal_map: dict[float, set[str]] = {}
    for i, d in enumerate(doses_raw):
        try:
            dose_val = float(d.strip())
            doses_set.add(dose_val)
            if i < len(animal_ids):
                if dose_val not in dose_animal_map:
                    dose_animal_map[dose_val] = set()
                dose_animal_map[dose_val].add(animal_ids[i])
        except (ValueError, TypeError):
            pass

    fp.dose_groups = sorted(doses_set)
    fp.n_animals_by_dose = {
        dose: len(aids) for dose, aids in sorted(dose_animal_map.items())
    }

    # Row 3+: endpoint names (first cell of each row)
    endpoint_names = []
    for line in lines[2:]:
        cells = line.split(separator)
        if cells and cells[0].strip():
            endpoint_names.append(cells[0].strip())
    fp.endpoint_names = endpoint_names

    # Gene expression detection — large files with many rows are probes/genes.
    # Only apply if filename-based detection didn't already assign a platform
    # (e.g., clinical_obs files can have 300+ observation types but aren't GE).
    if len(endpoint_names) > 200 and fp.platform is None:
        fp.data_type = "gene_expression"
        fp.gene_count = len(endpoint_names)
        # Extract organ from filename for gene expression files
        organ_match = _GENE_EXPR_ORGAN_PATTERN.match(filename)
        if organ_match:
            fp.organ = organ_match.group(1).title()

    return fp


def _extract_probe_ids(experiment: dict) -> list[str]:
    """
    Extract probe/endpoint IDs from a BMDExpress experiment.

    In the BMDProject JSON, probeResponses live on the experiment (not on
    each treatment).  Each probeResponse has a 'probe' field that is either:
      - A dict with an 'id' key (e.g., {"@type": "Probe", "id": "SD0"})
      - A stringified dict (Jackson serialization artifact)
      - An integer @ref (forward reference to another object)

    For apical endpoints, probe IDs are human-readable names like
    "Alanine aminotransferase", "SD0", "Heart", "Hematocrit".
    For gene expression, they're gene symbols like "AADAC_7934".

    Args:
        experiment: A doseResponseExperiment dict from the BMDProject JSON.

    Returns:
        List of probe ID strings.
    """
    probe_ids = []
    for pr in experiment.get("probeResponses", []):
        probe = pr.get("probe")
        if isinstance(probe, dict):
            probe_ids.append(str(probe.get("id", "")))
        elif isinstance(probe, str):
            # Jackson sometimes serializes nested objects as strings
            try:
                import ast
                probe_dict = ast.literal_eval(probe)
                if isinstance(probe_dict, dict):
                    probe_ids.append(str(probe_dict.get("id", "")))
                else:
                    probe_ids.append(probe)
            except (ValueError, SyntaxError):
                probe_ids.append(probe)
        elif isinstance(probe, int):
            # @ref — can't resolve without the full object graph
            probe_ids.append(f"@ref:{probe}")
    return probe_ids


def _extract_chip_info(experiment: dict) -> dict | None:
    """
    Extract chip/platform metadata from a BMDExpress experiment.

    The 'chip' field on doseResponseExperiment is either:
      - A dict with provider, species, name, geoName fields
        (e.g., {"provider": "BioSpyder", "species": "Rattus norvegicus",
                "name": "S1500_Plus_Rat"})
      - An integer @ref (when the same chip is shared across experiments)
      - None / null (for shoehorned clinical endpoint data)

    Returns None if chip info is unavailable or is just an @ref.

    Args:
        experiment: A doseResponseExperiment dict from the BMDProject JSON.

    Returns:
        Dict with provider, species, name keys, or None.
    """
    chip = experiment.get("chip")
    if isinstance(chip, dict):
        return {
            "provider": chip.get("provider"),
            "species": chip.get("species"),
            "name": chip.get("name") or chip.get("geoName"),
        }
    return None


def fingerprint_bm2(
    file_id: str,
    filename: str,
    path: str,
    ts_added: str,
    bm2_json: dict | None = None,
) -> FileFingerprint:
    """
    Extract metadata from a BMDExpress 3 .bm2 file's JSON representation.

    The BMDProject JSON (from Java export or LMDB cache) stores all data in
    a uniform schema regardless of data type.  The structure is:

      BMDProject
        doseResponseExperiments[]       — one per sex×domain (apical) or
                                          one per organ×sex (gene expression)
          name: str                     — e.g., "female_clin_chem", "Kidney_PFHxSAm_Male_No0"
          chip: dict|int|null           — platform info (BioSpyder, Affymetrix, or null)
          experimentDescription: dict|null — augmented metadata (null in current files)
          treatments[]                  — ONE PER ANIMAL (not per dose group)
            name: str                   — animal ID ("101", "Plate5-116")
            dose: float                 — dose assigned to this animal
          probeResponses[]              — the endpoints or genes
            probe: {id: str}            — endpoint name or gene probe ID
            responses: [float]          — one value per animal (same order as treatments)

    Clinical endpoints are "shoehorned" into this gene-expression schema:
      - "Alanine aminotransferase" is a "probe"
      - "SD0" (study day 0 body weight) is a "probe"
      - The category analysis uses "DEFINED-Category File-Non-genomic 5 day end points"
        where GO terms would normally go

    This function extracts the actual structural data (doses, animal IDs,
    endpoint names from probe.id) and uses an LLM call (Haiku) to infer
    semantic metadata (sex, organ, species, strain, domain) from the
    experiment names — the same approach BMDExpress-3's own
    LlmMetadataService uses on the Java side.

    If bm2_json is not provided, attempts to load from LMDB cache.
    Does NOT trigger Java export (expensive) — returns a partial
    fingerprint if the JSON is unavailable.

    Args:
        file_id:    UUID from upload.
        filename:   Original filename as uploaded.
        path:       Absolute path to the .bm2 file on disk.
        ts_added:   ISO 8601 timestamp when the file entered the pool.
        bm2_json:   Pre-loaded BMDProject dict (optional — avoids cache lookup).

    Returns:
        FileFingerprint with domain, sexes, doses, endpoints, animal IDs, and
        species/strain metadata inferred by LLM.
    """
    fp = FileFingerprint(
        file_id=file_id,
        filename=filename,
        file_type="bm2",
        tier=TIER_BM2,
        ts_added=ts_added,
        ts_filesystem=_get_filesystem_timestamp(path),
    )

    # Filename-based platform/type detection as a fallback
    fp.platform, fp.data_type = detect_platform_and_type(filename, "bm2")

    # Try loading from LMDB cache if not provided — this is fast (~10ms)
    # and avoids the expensive Java export on fingerprinting.
    if bm2_json is None:
        try:
            import bm2_cache
            bm2_json = bm2_cache.get_json(path)
        except Exception:
            pass

    if bm2_json is None:
        # No JSON available — return partial fingerprint.  The domain was
        # already detected from filename; we just can't extract dose groups
        # or experiment names without the JSON.
        logger.debug("No cached JSON for %s — partial fingerprint only", filename)
        return fp

    # --- Extract structural data from doseResponseExperiments ---
    experiments = bm2_json.get("doseResponseExperiments", [])
    if not experiments:
        return fp

    experiment_names = [exp.get("name", "") for exp in experiments]

    # --- Extract probe/endpoint IDs from all experiments ---
    # For apical files, these are the actual clinical endpoint names
    # (e.g., "Alanine aminotransferase", "SD0", "Heart").
    # For gene expression, these are gene probe IDs (e.g., "AADAC_7934").
    all_probe_ids: list[str] = []
    for exp in experiments:
        all_probe_ids.extend(_extract_probe_ids(exp))
    # Deduplicate while preserving order (probes are shared across experiments)
    seen_probes: set[str] = set()
    unique_probes: list[str] = []
    for pid in all_probe_ids:
        if pid not in seen_probes:
            seen_probes.add(pid)
            unique_probes.append(pid)

    # --- Extract treatments (animal IDs and doses) from first experiment ---
    # Treatments are per-animal, not per-dose-group.  Each treatment has a
    # 'name' (animal ID like "101") and a 'dose' (concentration assigned).
    first_exp = experiments[0]
    treatments = first_exp.get("treatments", [])

    doses_set: set[float] = set()
    animal_ids: list[str] = []
    animals_per_dose: dict[float, set[str]] = {}

    for treatment in treatments:
        dose = treatment.get("dose")
        animal_name = treatment.get("name", "")
        if dose is not None:
            try:
                dose_val = float(dose)
                doses_set.add(dose_val)
                if animal_name:
                    if dose_val not in animals_per_dose:
                        animals_per_dose[dose_val] = set()
                    animals_per_dose[dose_val].add(animal_name)
            except (ValueError, TypeError):
                pass
        if animal_name:
            animal_ids.append(animal_name)

    fp.dose_groups = sorted(doses_set)
    fp.animal_ids = sorted(set(animal_ids))
    fp.n_animals_by_dose = {
        dose: len(aids) for dose, aids in sorted(animals_per_dose.items())
    }

    # --- Extract chip/platform info ---
    chip_info = _extract_chip_info(first_exp)
    if chip_info and chip_info.get("species"):
        fp.species = chip_info["species"]

    # --- Check for augmented ExperimentDescription (future bm2 files) ---
    # When present, this provides authoritative metadata that supersedes
    # LLM deduction.  Current files have experimentDescription: null.
    exp_desc = first_exp.get("experimentDescription")
    if isinstance(exp_desc, dict) and exp_desc:
        # Future augmented schema — extract directly
        if exp_desc.get("sex"):
            fp.sexes = [exp_desc["sex"].title()]
        if exp_desc.get("organ"):
            fp.organ = exp_desc["organ"].title()
        if exp_desc.get("species"):
            fp.species = exp_desc["species"]
        # Set platform directly from augmented ExperimentDescription.
        # .bm2 files are always produced from gap-filled (inferred) data,
        # so data_type is always "inferred" for apical bm2.
        desc_platform = exp_desc.get("platform", "")
        if desc_platform:
            fp.platform = desc_platform
            fp.data_type = "inferred"
        provider = exp_desc.get("provider", "")
        if provider == "Clinical Endpoint" and not fp.platform:
            fp.platform = "Clinical"
            fp.data_type = "inferred"
        logger.debug("Using augmented ExperimentDescription for %s", filename)
    else:
        # --- LLM-based metadata deduction ---
        # Current .bm2 files lack ExperimentDescription, so we ask an LLM
        # to infer sex, organ, species, strain, and platform from the
        # experiment names and probe ID samples.  This mirrors the approach
        # in BMDExpress-3's LlmMetadataService.
        probe_sample = unique_probes[:15]
        llm_meta = _deduce_metadata_from_experiments(
            experiment_names, probe_sample, chip_info,
        )

        if llm_meta:
            # Apply LLM-inferred metadata, falling back to regex on miss
            if llm_meta.get("sexes"):
                fp.sexes = sorted(
                    s.title() for s in llm_meta["sexes"]
                    if isinstance(s, str)
                )
            # LLM returns "platform" (new) or "domain" (old prompt format).
            # Handle both for transition period.
            llm_platform = llm_meta.get("platform") or llm_meta.get("domain")
            if llm_platform:
                # LLM may return a list — take the first string
                if isinstance(llm_platform, list):
                    llm_platform = llm_platform[0] if llm_platform else None
                if isinstance(llm_platform, str):
                    # LLM might return old-style domain or new platform name.
                    # Check if it's a known platform value first.
                    _KNOWN_PLATFORMS = {
                        "Body Weight", "Organ Weight", "Clinical Chemistry",
                        "Hematology", "Hormones", "Tissue Concentration",
                        "Clinical",
                    }
                    if llm_platform in _KNOWN_PLATFORMS:
                        fp.platform = llm_platform
                        fp.data_type = "inferred"
                    elif llm_platform == "gene_expression":
                        fp.platform = None
                        fp.data_type = "gene_expression"
                    else:
                        # Old-style domain string from LLM — try to parse it
                        # by running it through the old detect_domain path
                        # and extracting the platform from that.
                        old_platform, old_dtype = detect_platform_and_type(
                            llm_platform, "bm2",
                        )
                        if old_platform:
                            fp.platform = old_platform
                            fp.data_type = old_dtype or "inferred"
            if llm_meta.get("species") and not fp.species:
                # Ignore placeholder values like "generic" that Haiku returns
                # when species can't be determined from experiment names alone
                if llm_meta["species"].lower() not in ("generic", "unknown", "na", "null"):
                    fp.species = llm_meta["species"]
            if llm_meta.get("strain"):
                fp.species = (
                    f"{llm_meta['strain']} "
                    f"({llm_meta.get('species', '')})"
                ).strip(" ()")
                # Store strain info in species field since we don't have
                # a separate strain field on FileFingerprint
                if llm_meta.get("species"):
                    fp.species = f"{llm_meta['strain']} {llm_meta['species']}"
            if llm_meta.get("organs"):
                organs = [o.title() for o in llm_meta["organs"] if isinstance(o, str)]
                # Only set organ for gene expression — apical platforms like
                # Organ Weight and Hematology measure *across* organs, so
                # organ is semantically wrong there.
                if organs and fp.data_type == "gene_expression":
                    fp.organ = organs[0]  # primary organ for transcriptomics

        # Regex fallback if LLM didn't return a platform
        if not fp.platform and fp.data_type != "gene_expression":
            fp.platform, fp.data_type = detect_platform_and_type_from_bm2(
                experiment_names,
            )
        # Regex fallback for sex detection
        if not fp.sexes:
            fp.sexes = _detect_sexes_from_bm2_experiments(experiment_names)

    # --- Set endpoint names and gene count ---
    if fp.data_type == "gene_expression":
        # For gene expression, probeResponses are genes — count them per
        # experiment (each experiment = one organ×sex combination).
        # The total gene count is from the first experiment.
        fp.gene_count = len(first_exp.get("probeResponses", []))
        fp.endpoint_names = experiment_names  # organ×sex experiments
    else:
        # For apical endpoints, the probe IDs are the actual endpoint names
        # (e.g., "Albumin", "SD0", "Heart").  These are what appear as
        # row labels in the NTP tables.
        fp.endpoint_names = unique_probes

    # --- Study metadata from project-level fields ---
    for key in ("name", "projectName", "studyName"):
        val = bm2_json.get(key, "")
        if val and re.match(r"^[A-Z]\d+-\d+$", str(val)):
            fp.study_number = str(val)
            break

    return fp


def fingerprint_file(
    file_id: str,
    filename: str,
    path: str,
    file_type: str,
    ts_added: str,
    bm2_json: dict | None = None,
) -> FileFingerprint:
    """
    Router — dispatches to the appropriate fingerprint extractor by file_type.

    This is the main entry point for fingerprinting.  Call this instead of
    the type-specific functions to ensure consistent handling.

    Args:
        file_id:    UUID from upload.
        filename:   Original filename as uploaded.
        path:       Absolute path to the file on disk.
        file_type:  "xlsx", "txt", "csv", or "bm2".
        ts_added:   ISO 8601 timestamp when the file entered the pool.
        bm2_json:   Pre-loaded BMDProject dict (only used for bm2 files).

    Returns:
        FileFingerprint with as much metadata as could be extracted.
    """
    if file_type == "xlsx":
        return fingerprint_xlsx(file_id, filename, path, ts_added)
    elif file_type in ("txt", "csv"):
        return fingerprint_txt_csv(file_id, filename, path, file_type, ts_added)
    elif file_type == "bm2":
        return fingerprint_bm2(file_id, filename, path, ts_added, bm2_json)
    else:
        # Unknown file type — return a minimal fingerprint
        logger.warning("Unknown file type '%s' for %s", file_type, filename)
        return FileFingerprint(
            file_id=file_id,
            filename=filename,
            file_type=file_type,
            tier=99,
            ts_added=ts_added,
            ts_filesystem=_get_filesystem_timestamp(path),
        )


# ---------------------------------------------------------------------------
# Cross-validation engine
# ---------------------------------------------------------------------------


def _build_coverage_matrix(
    fingerprints: dict[str, FileFingerprint],
) -> dict[str, dict[str, list[str] | str | None]]:
    """
    Build a coverage matrix mapping (platform, data_type) → tier → file_id(s).

    Groups fingerprints by platform AND data_type so that tox_study and
    inferred files for the same platform get separate coverage entries.
    This ensures both are included in integration — tox_study data provides
    raw values for NTP stats, inferred .bm2 data provides BMD results.

    The compound key format is "Platform|data_type" (e.g., "Body Weight|tox_study",
    "Body Weight|inferred").  Gene expression uses "gene_expression".

    Args:
        fingerprints: All fingerprints in the pool (file_id → FileFingerprint).

    Returns:
        Dict of compound_key → { "xlsx": file_id|None, "txt_csv": [file_ids], "bm2": file_id|None }
    """
    matrix: dict[str, dict] = {}

    for fid, fp in fingerprints.items():
        # Determine the grouping key — compound of platform + data_type.
        # Gene expression (platform=None) uses just "gene_expression".
        if fp.data_type == "gene_expression":
            group_key = "gene_expression"
        elif fp.platform is not None and fp.data_type:
            group_key = f"{fp.platform}|{fp.data_type}"
        elif fp.platform is not None:
            group_key = fp.platform
        else:
            continue  # unrecognized file — skip

        if group_key not in matrix:
            matrix[group_key] = {"xlsx": None, "txt_csv": [], "bm2": None}

        entry = matrix[group_key]
        if fp.tier == TIER_XLSX:
            entry["xlsx"] = fid
        elif fp.tier == TIER_TXT_CSV:
            entry["txt_csv"].append(fid)
        elif fp.tier == TIER_BM2:
            entry["bm2"] = fid

    return matrix


def _check_coverage(
    domain: str,
    tiers: dict,
    fingerprints: dict[str, FileFingerprint],
) -> list[ValidationIssue]:
    """
    Check tier coverage for a single domain.

    Reports missing tiers as warnings or info depending on the domain.
    Gene expression typically has no xlsx (transcriptomics data comes as
    txt from the microarray platform, not from NTP study team xlsx).

    Args:
        domain:       The endpoint domain (e.g., "body_weight").
        tiers:        The coverage entry { "xlsx": ..., "txt_csv": [...], "bm2": ... }
        fingerprints: All fingerprints in the pool.

    Returns:
        List of ValidationIssue objects for missing tiers.
    """
    issues = []

    has_xlsx = tiers["xlsx"] is not None
    has_txt_csv = len(tiers["txt_csv"]) > 0
    has_bm2 = tiers["bm2"] is not None

    # Gene expression: xlsx is not expected (data comes as txt from microarray)
    if domain == "gene_expression":
        if not has_txt_csv:
            issues.append(ValidationIssue(
                severity="warning",
                platform=domain,
                issue_type="missing_tier",
                message=f"{domain}: no txt/csv gene expression input data found",
                files_involved=[tiers["bm2"]] if has_bm2 else [],
            ))
        if not has_bm2 and has_txt_csv:
            issues.append(ValidationIssue(
                severity="warning",
                platform=domain,
                issue_type="missing_tier",
                message=f"{domain}: txt/csv exists but no .bm2 analysis found",
                files_involved=tiers["txt_csv"],
            ))
        return issues

    # Apical platforms: all three tiers are expected (xlsx → txt/csv → bm2)
    if not has_xlsx:
        issues.append(ValidationIssue(
            severity="info",
            platform=domain,
            issue_type="missing_tier",
            message=f"{domain}: no xlsx (study team raw data) found",
            files_involved=tiers["txt_csv"] + ([tiers["bm2"]] if has_bm2 else []),
        ))

    if not has_txt_csv:
        # Missing txt/csv means can't re-run BMDExpress — significant gap
        issues.append(ValidationIssue(
            severity="warning",
            platform=domain,
            issue_type="missing_tier",
            message=f"{domain}: no txt/csv (BMDExpress input) found",
            files_involved=([tiers["xlsx"]] if has_xlsx else [])
                         + ([tiers["bm2"]] if has_bm2 else []),
        ))

    if not has_bm2:
        # Missing bm2 means analysis not yet done for this platform
        sev = "warning" if (has_xlsx or has_txt_csv) else "info"
        issues.append(ValidationIssue(
            severity=sev,
            platform=domain,
            issue_type="missing_tier",
            message=f"{domain}: no .bm2 (BMDExpress analysis) found",
            files_involved=([tiers["xlsx"]] if has_xlsx else [])
                         + tiers["txt_csv"],
        ))

    return issues


def _check_dose_consistency(
    domain: str,
    tiers: dict,
    fingerprints: dict[str, FileFingerprint],
) -> list[ValidationIssue]:
    """
    Compare dose_groups across all files in the same domain.

    Dose groups must match exactly between xlsx and txt/csv (txt/csv is
    derived from xlsx), and between txt/csv and bm2 (bm2 is derived from
    txt/csv).  Any mismatch is an error requiring user resolution.

    Args:
        domain:       The endpoint domain.
        tiers:        The coverage entry from the coverage matrix.
        fingerprints: All fingerprints in the pool.

    Returns:
        List of ValidationIssue objects for dose group mismatches.
    """
    issues = []

    # Collect all dose group sets in this domain, tagged by file
    dose_sets: list[tuple[str, list[float]]] = []

    if tiers["xlsx"]:
        fp = fingerprints[tiers["xlsx"]]
        if fp.dose_groups:
            dose_sets.append((fp.file_id, fp.dose_groups))

    for txt_id in tiers["txt_csv"]:
        fp = fingerprints[txt_id]
        if fp.dose_groups:
            dose_sets.append((fp.file_id, fp.dose_groups))

    if tiers["bm2"]:
        fp = fingerprints[tiers["bm2"]]
        if fp.dose_groups:
            dose_sets.append((fp.file_id, fp.dose_groups))

    if len(dose_sets) < 2:
        return issues  # Can't compare with fewer than 2

    # Compare all pairs — O(n²) but n is tiny (at most ~4 files per domain).
    # Only compare files that share the same sex.  Male and female files
    # within the same domain are independent — different animals, possibly
    # different dose groups (high-dose deaths may differ by sex).
    # Files with no detected sex are compared against everything (conservative).
    def _sexes_overlap(fp_a: FileFingerprint, fp_b: FileFingerprint) -> bool:
        """True if two files should be compared (same sex or sex unknown)."""
        if not fp_a.sexes or not fp_b.sexes:
            return True  # unknown sex → compare conservatively
        return bool(set(fp_a.sexes) & set(fp_b.sexes))

    for i, (ref_id, ref_doses) in enumerate(dose_sets):
        for other_id, other_doses in dose_sets[i + 1:]:
            ref_fp = fingerprints[ref_id]
            other_fp = fingerprints[other_id]

            # Skip comparison if files have different data types —
            # tox_study files (raw, with gaps) are expected to have more
            # dose groups than inferred files (high-dose groups excluded
            # when all animals died).
            if ref_fp.data_type and other_fp.data_type and ref_fp.data_type != other_fp.data_type:
                continue

            # Skip comparison if files are for different sexes
            if not _sexes_overlap(ref_fp, other_fp):
                continue

            if ref_doses != other_doses:
                suggested = ref_id if ref_fp.tier <= other_fp.tier else other_id
                issues.append(ValidationIssue(
                    severity="error",
                    platform=domain,
                    issue_type="dose_mismatch",
                    message=(
                        f"{domain}: dose groups differ between "
                        f"{ref_fp.filename} ({ref_fp.file_type}) and "
                        f"{other_fp.filename} ({other_fp.file_type})"
                    ),
                    files_involved=[ref_id, other_id],
                    suggested_precedence=suggested,
                    details={
                        "expected": ref_doses,
                        "actual": other_doses,
                        "expected_file": ref_fp.filename,
                        "actual_file": other_fp.filename,
                    },
                ))

    return issues


def _check_animal_counts(
    domain: str,
    tiers: dict,
    fingerprints: dict[str, FileFingerprint],
) -> list[ValidationIssue]:
    """
    Compare animal counts per dose group across tiers.

    txt/csv files are split by sex (e.g., male_body_weight.txt has only male
    animals), while xlsx has both sexes.  So we compare the *sum* of txt/csv
    animal counts against the xlsx per-sex counts.

    For now, we just compare total unique animal counts between xlsx and
    the union of all txt/csv files in the same domain.

    Args:
        domain:       The endpoint domain.
        tiers:        The coverage entry from the coverage matrix.
        fingerprints: All fingerprints in the pool.

    Returns:
        List of ValidationIssue objects for animal count mismatches.
    """
    issues = []

    xlsx_id = tiers["xlsx"]
    txt_csv_ids = tiers["txt_csv"]

    if not xlsx_id or not txt_csv_ids:
        return issues  # Need both tiers to compare

    xlsx_fp = fingerprints[xlsx_id]
    xlsx_animal_count = len(xlsx_fp.animal_ids)

    # Union of all txt/csv animal IDs in this domain
    txt_csv_animals: set[str] = set()
    for tid in txt_csv_ids:
        tfp = fingerprints[tid]
        txt_csv_animals.update(tfp.animal_ids)
    txt_csv_count = len(txt_csv_animals)

    if xlsx_animal_count > 0 and txt_csv_count > 0:
        if txt_csv_count > xlsx_animal_count:
            # Extra animals in txt/csv — something is wrong
            issues.append(ValidationIssue(
                severity="error",
                platform=domain,
                issue_type="animal_count_mismatch",
                message=(
                    f"{domain}: txt/csv files have {txt_csv_count} unique animals "
                    f"but xlsx has only {xlsx_animal_count} "
                    f"({txt_csv_count - xlsx_animal_count} extra)"
                ),
                files_involved=[xlsx_id] + txt_csv_ids,
                suggested_precedence=xlsx_id,
                details={
                    "xlsx_count": xlsx_animal_count,
                    "txt_csv_count": txt_csv_count,
                },
            ))
        elif txt_csv_count < xlsx_animal_count:
            # Fewer animals in txt/csv — some may have been excluded (common)
            issues.append(ValidationIssue(
                severity="warning",
                platform=domain,
                issue_type="animal_count_mismatch",
                message=(
                    f"{domain}: txt/csv files have {txt_csv_count} unique animals "
                    f"vs {xlsx_animal_count} in xlsx "
                    f"({xlsx_animal_count - txt_csv_count} excluded)"
                ),
                files_involved=[xlsx_id] + txt_csv_ids,
                suggested_precedence=xlsx_id,
                details={
                    "xlsx_count": xlsx_animal_count,
                    "txt_csv_count": txt_csv_count,
                },
            ))

    return issues


def _check_sex_coverage(
    domain: str,
    tiers: dict,
    fingerprints: dict[str, FileFingerprint],
) -> list[ValidationIssue]:
    """
    Ensure male/female split is consistent across tiers.

    xlsx files have both sexes in one file; txt/csv files are split by sex
    (one file per sex per domain).  bm2 experiments encode sex in the name.

    Checks that the set of sexes in txt/csv matches the xlsx and bm2.

    Args:
        domain:       The endpoint domain.
        tiers:        The coverage entry from the coverage matrix.
        fingerprints: All fingerprints in the pool.

    Returns:
        List of ValidationIssue objects for sex coverage mismatches.
    """
    issues = []

    # Collect sexes from each tier
    xlsx_sexes: set[str] = set()
    txt_csv_sexes: set[str] = set()
    bm2_sexes: set[str] = set()

    if tiers["xlsx"]:
        xlsx_sexes = set(fingerprints[tiers["xlsx"]].sexes)
    for tid in tiers["txt_csv"]:
        txt_csv_sexes.update(fingerprints[tid].sexes)
    if tiers["bm2"]:
        bm2_sexes = set(fingerprints[tiers["bm2"]].sexes)

    # Compare xlsx vs txt_csv
    if xlsx_sexes and txt_csv_sexes and xlsx_sexes != txt_csv_sexes:
        missing_in_txt = xlsx_sexes - txt_csv_sexes
        if missing_in_txt:
            issues.append(ValidationIssue(
                severity="warning",
                platform=domain,
                issue_type="sex_mismatch",
                message=(
                    f"{domain}: xlsx has {sorted(xlsx_sexes)} but txt/csv "
                    f"only covers {sorted(txt_csv_sexes)} "
                    f"(missing: {sorted(missing_in_txt)})"
                ),
                files_involved=[tiers["xlsx"]] + tiers["txt_csv"],
            ))

    # Compare txt_csv vs bm2
    if txt_csv_sexes and bm2_sexes and txt_csv_sexes != bm2_sexes:
        missing_in_bm2 = txt_csv_sexes - bm2_sexes
        if missing_in_bm2:
            issues.append(ValidationIssue(
                severity="warning",
                platform=domain,
                issue_type="sex_mismatch",
                message=(
                    f"{domain}: txt/csv covers {sorted(txt_csv_sexes)} but "
                    f".bm2 only has {sorted(bm2_sexes)} "
                    f"(missing: {sorted(missing_in_bm2)})"
                ),
                files_involved=tiers["txt_csv"] + ([tiers["bm2"]] if tiers["bm2"] else []),
            ))

    return issues


def _check_redundancy(
    fingerprints: dict[str, FileFingerprint],
) -> list[ValidationIssue]:
    """
    Detect files with identical structural fingerprints (same data, different formats).

    Two files are considered redundant if they share the same domain, sexes,
    dose_groups, and endpoint_names — meaning they contain the same data in
    different file formats (e.g., male_body_weight.csv and male_body_weight.txt).

    Args:
        fingerprints: All fingerprints in the pool.

    Returns:
        List of ValidationIssue objects for redundant file pairs.
    """
    issues = []
    seen: dict[tuple, str] = {}  # signature → first file_id

    for fid, fp in fingerprints.items():
        # Build a structural signature: (platform, data_type, sorted_sexes,
        # sorted_doses, sorted_endpoints, animal_id_sample).  Including
        # animal_ids prevents false positives for sex-split files (male and
        # female txt files share the same platform/endpoints but have
        # different animal cohorts).
        sig = (
            fp.platform,
            fp.data_type,
            tuple(fp.sexes),
            tuple(fp.dose_groups),
            tuple(fp.endpoint_names[:20]),  # truncate for efficiency
            tuple(sorted(fp.animal_ids)[:10]),  # sample of animal IDs
        )
        if sig in seen:
            other_fid = seen[sig]
            other_fp = fingerprints[other_fid]
            # Only flag as redundant if they're the same tier (different tiers
            # are expected to have the same data at different processing stages)
            if fp.tier == other_fp.tier:
                issues.append(ValidationIssue(
                    severity="info",
                    platform=fp.platform or "unknown",
                    issue_type="redundant_files",
                    message=(
                        f"{fp.filename} and {other_fp.filename} appear to "
                        f"contain the same data (same platform, sexes, doses, endpoints)"
                    ),
                    files_involved=[fid, other_fid],
                ))
        else:
            seen[sig] = fid

    return issues


def _check_roster_consistency(
    fingerprints: dict[str, FileFingerprint],
) -> list[ValidationIssue]:
    """
    Top-level roster validation — runs before all other checks.

    The study roster per sex is the UNION of animal IDs from ALL files
    (all domains, all tiers).  No domain has special roster authority —
    every file that reports animal IDs contributes to the universal roster.

    Once the universal roster is built, each individual file's animal set
    must be a subset.  An animal ID appearing in only one file is fine
    (that file simply has the most complete roster).  An animal ID NOT
    in the universal set is impossible by construction (union includes
    everything), so the only check is whether files are internally
    consistent (no IDs that appear in one file but are clearly erroneous).

    Currently this is a no-op warning generator: it reports which files
    have fewer animals than the largest file for each sex, as an FYI.
    This helps users spot files that are missing animals (dead animals,
    excluded dose groups) without blocking processing.

    Args:
        fingerprints: Dict of file_id → FileFingerprint for all pool files.

    Returns:
        List of ValidationIssue objects (informational, not blocking).
    """
    issues: list[ValidationIssue] = []

    # --- Build universal roster per sex from ALL files ---
    # files_by_sex: {"Male": {file_id: set(animal_ids), ...}, ...}
    files_by_sex: dict[str, dict[str, set[str]]] = {}

    for fid, fp in fingerprints.items():
        if not fp.animal_ids:
            continue
        # Gene expression files use plate/sample IDs (e.g., "Plate5-116"),
        # not study animal IDs (e.g., "101").  Exclude from roster.
        if fp.data_type == "gene_expression":
            continue

        sexes = fp.sexes if fp.sexes else ["Unknown"]
        for sex in sexes:
            files_by_sex.setdefault(sex, {})[fid] = set(fp.animal_ids)

    # For each sex, compute the universal roster (union) and report
    # files that have fewer animals as informational warnings.
    for sex, files in files_by_sex.items():
        if not files:
            continue

        # Universal roster = union of all animal IDs for this sex
        roster: set[str] = set()
        for aids in files.values():
            roster |= aids

        # Find the file with the most animals (typically body weight)
        largest_fid = max(files, key=lambda f: len(files[f]))
        largest_fp = fingerprints[largest_fid]

        for fid, file_animals in files.items():
            missing = roster - file_animals
            if missing and len(missing) <= len(roster) * 0.5:
                # File has fewer animals than the universal roster —
                # expected for domains where high-dose animals died.
                fp = fingerprints[fid]
                issues.append(ValidationIssue(
                    severity="info",
                    platform=fp.platform or "unknown",
                    issue_type="roster_subset",
                    message=(
                        f"{fp.filename} ({sex}): {len(file_animals)}/{len(roster)} "
                        f"animals — missing {len(missing)} "
                        f"({', '.join(sorted(list(missing)[:5]))})"
                    ),
                    files_involved=[fid],
                    details={
                        "sex": sex,
                        "file_count": len(file_animals),
                        "roster_count": len(roster),
                        "missing_count": len(missing),
                    },
                ))

    return issues


def validate_pool(
    dtxsid: str,
    fingerprints: dict[str, FileFingerprint],
) -> ValidationReport:
    """
    Run full cross-validation on a session's file pool.

    Validation hierarchy (each level blocks the next):
      0. Roster consistency — _tox_study files define the authoritative animal
         roster per sex.  All _tox_study files for the same sex must agree.
         Every other file must be a subset.  Errors here block everything.
      1. Domain assignment — every file must have a recognized domain.
         Unassigned files block all processing.
      2. Intra-domain dose consistency — files within the same domain + sex
         must agree on dose groups.
      3. Animal count consistency — do animal counts match across tiers?
      4. Sex coverage — is the male/female split consistent?
      5. Redundancy detection — are there duplicate files?

    Builds a coverage matrix showing the tier completeness per domain.

    Args:
        dtxsid:       The DTXSID identifier for this session.
        fingerprints: Dict of file_id → FileFingerprint for all pool files.

    Returns:
        ValidationReport with issues, coverage matrix, and completeness flag.
    """
    now = datetime.now(tz=timezone.utc).isoformat()
    matrix = _build_coverage_matrix(fingerprints)
    all_issues: list[ValidationIssue] = []

    # --- Level 0: Roster consistency (top-level, blocks everything) ---
    roster_issues = _check_roster_consistency(fingerprints)
    all_issues.extend(roster_issues)

    # --- Level 1: Unassigned files block all processing ---
    # A file is unassigned if it has no platform AND no data_type — meaning
    # neither filename patterns nor content-based detection could identify it.
    for fid, fp in fingerprints.items():
        if fp.platform is None and fp.data_type is None:
            all_issues.append(ValidationIssue(
                severity="error",
                platform="unassigned",
                issue_type="unassigned_file",
                message=(
                    f"{fp.filename} could not be assigned to a recognized platform. "
                    f"Remove it or rename it so the system can identify its purpose."
                ),
                files_involved=[fid],
            ))

    # --- Level 2+: Per-domain checks ---
    for domain, tiers in matrix.items():
        all_issues.extend(_check_coverage(domain, tiers, fingerprints))
        all_issues.extend(_check_dose_consistency(domain, tiers, fingerprints))
        all_issues.extend(_check_animal_counts(domain, tiers, fingerprints))
        all_issues.extend(_check_sex_coverage(domain, tiers, fingerprints))

    # Cross-domain checks
    all_issues.extend(_check_redundancy(fingerprints))

    # Compute completeness — every domain with any file should have all
    # applicable tiers.  Gene expression doesn't need xlsx.
    is_complete = True
    for domain, tiers in matrix.items():
        if domain == "gene_expression":
            if not tiers["txt_csv"] or tiers["bm2"] is None:
                is_complete = False
        else:
            if tiers["xlsx"] is None or not tiers["txt_csv"] or tiers["bm2"] is None:
                is_complete = False

    # Serialize coverage matrix for JSON output — convert lists to
    # JSON-friendly format
    coverage_json = {}
    for domain, tiers in matrix.items():
        coverage_json[domain] = {
            "xlsx": tiers["xlsx"],
            "txt_csv": tiers["txt_csv"],
            "bm2": tiers["bm2"],
        }

    # Serialize fingerprints with JSON-safe keys — several dict fields use
    # float dose keys which orjson rejects (dict keys must be str).  Convert
    # them to strings here for safe serialization.
    fp_dicts = {}
    for fid, fp in fingerprints.items():
        d = asdict(fp)
        for float_key_field in (
            "n_animals_by_dose",
            "animals_by_dose_selection",
            "core_animals_by_dose_sex",
        ):
            if d.get(float_key_field):
                d[float_key_field] = {
                    str(k): v for k, v in d[float_key_field].items()
                }
        fp_dicts[fid] = d

    return ValidationReport(
        dtxsid=dtxsid,
        run_at=now,
        file_count=len(fingerprints),
        fingerprints=fp_dicts,
        issues=[asdict(issue) for issue in all_issues],
        coverage_matrix=coverage_json,
        is_complete=is_complete,
    )


def lightweight_validate(
    new_fp: FileFingerprint,
    existing_fps: dict[str, FileFingerprint],
) -> list[ValidationIssue]:
    """
    Quick validation of a single new file against the existing pool.

    Called on file addition (upload or zip extraction) to provide immediate
    feedback without running full pool validation.  Checks:
      1. Dose group match against existing files in the same domain
      2. Obvious redundancy (same domain + tier + endpoints)

    Much faster than full validation (~10ms) because it only examines
    files in the same domain as the new file.

    Args:
        new_fp:       Fingerprint of the newly added file.
        existing_fps: Existing fingerprints in the pool (file_id → fingerprint).

    Returns:
        List of ValidationIssue objects (may be empty if no issues).
    """
    issues: list[ValidationIssue] = []

    # A file needs either a platform or a data_type to be matchable
    if new_fp.platform is None and new_fp.data_type is None:
        return issues

    # Determine the grouping key for matching — same logic as _build_coverage_matrix
    if new_fp.data_type == "gene_expression":
        new_group_key = "gene_expression"
    else:
        new_group_key = new_fp.platform

    if new_group_key is None:
        return issues

    # Find existing files in the same platform group
    def _group_key(fp: FileFingerprint) -> str | None:
        if fp.data_type == "gene_expression":
            return "gene_expression"
        return fp.platform

    same_platform = {
        fid: fp for fid, fp in existing_fps.items()
        if _group_key(fp) == new_group_key
    }

    if not same_platform:
        return issues

    # Check dose group consistency
    if new_fp.dose_groups:
        for fid, fp in same_platform.items():
            if fp.dose_groups and fp.dose_groups != new_fp.dose_groups:
                suggested = fid if fp.tier < new_fp.tier else new_fp.file_id
                issues.append(ValidationIssue(
                    severity="error",
                    platform=new_group_key,
                    issue_type="dose_mismatch",
                    message=(
                        f"Dose groups in {new_fp.filename} differ from "
                        f"{fp.filename} ({fp.file_type})"
                    ),
                    files_involved=[new_fp.file_id, fid],
                    suggested_precedence=suggested,
                    details={
                        "new_doses": new_fp.dose_groups,
                        "existing_doses": fp.dose_groups,
                    },
                ))

    # Check for redundancy (same tier + similar structure)
    new_sig = (new_fp.platform, new_fp.data_type, tuple(new_fp.sexes), tuple(new_fp.dose_groups))
    for fid, fp in same_platform.items():
        if fp.tier == new_fp.tier:
            existing_sig = (fp.platform, fp.data_type, tuple(fp.sexes), tuple(fp.dose_groups))
            if new_sig == existing_sig and new_fp.endpoint_names[:10] == fp.endpoint_names[:10]:
                issues.append(ValidationIssue(
                    severity="info",
                    platform=new_group_key,
                    issue_type="redundant_files",
                    message=(
                        f"{new_fp.filename} appears redundant with "
                        f"{fp.filename} (same platform, sexes, doses)"
                    ),
                    files_involved=[new_fp.file_id, fid],
                ))

    return issues
